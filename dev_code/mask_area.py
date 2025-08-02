'''
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract plant shoot traits (larea, solidity, max_width, max_height, avg_curv, color_cluster) by paralell processing 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2024-02-29

USAGE:

    python3 mask_area.py -p /input/ -o /output/

'''

# import the necessary packages
import subprocess, os, glob, sys
import utils

from collections import Counter
from collections import OrderedDict

import argparse

from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from skimage.feature import peak_local_max
from skimage.morphology import medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure
from skimage.color import rgb2lab, deltaE_cie76
from skimage import morphology
from skimage.segmentation import clear_border, watershed
from skimage.measure import regionprops

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage
from scipy.interpolate import interp1d

#from skan import skeleton_to_csgraph, Skeleton, summarize, draw

#import networkx as nx

import imutils

import numpy as np
import argparse
import cv2

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import math
import openpyxl
import csv
    
from tabulate import tabulate

import warnings
warnings.filterwarnings("ignore")

import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from pathlib import Path

from rembg import remove

from matplotlib import collections

import matplotlib.colors

import pathlib


MBFACTOR = float(1<<20)


# check file type
def check_file_type(image_folder_path, allowed_extensions=None):
    
    if allowed_extensions is None:
        allowed_extensions = ['.jpg', '.png', '.jpeg']

    no_files_in_folder = len(glob.glob(image_folder_path+"/*")) 
    extension_type = ""
    no_files_allowed = 0

    for ext in allowed_extensions:
      no_files_allowed = len(glob.glob(image_folder_path+"/*"+ext))
      if no_files_allowed > 0:
        extension_type = ext

    return extension_type


# curvature computation calss
class ComputeCurvature:

    def __init__(self,x,y):
        """ Initialize some variables """
        self.xc = 0  # X-coordinate of circle center
        self.yc = 0  # Y-coordinate of circle center
        self.r = 0   # Radius of the circle
        self.xx = np.array([])  # Data points
        self.yy = np.array([])  # Data points
        self.x = x  # X-coordinate of circle center
        self.y = y  # Y-coordinate of circle center

    def calc_r(self, xc, yc):
        """ calculate the distance of each 2D points from the center (xc, yc) """
        return np.sqrt((self.xx-xc)**2 + (self.yy-yc)**2)

    def f(self, c):
        """ calculate the algebraic distance between the data points and the mean circle centered at c=(xc, yc) """
        ri = self.calc_r(*c)
        return ri - ri.mean()

    def df(self, c):
        """ Jacobian of f_2b
        The axis corresponding to derivatives must be coherent with the col_deriv option of leastsq"""
        xc, yc = c
        df_dc = np.empty((len(c), self.x.size))

        ri = self.calc_r(xc, yc)
        df_dc[0] = (xc - self.x)/ri                   # dR/dxc
        df_dc[1] = (yc - self.y)/ri                   # dR/dyc
        df_dc = df_dc - df_dc.mean(axis=1)[:, np.newaxis]
        return df_dc

    def fit(self, xx, yy):
        self.xx = xx
        self.yy = yy
        center_estimate = np.r_[np.mean(xx), np.mean(yy)]
        center = optimize.leastsq(self.f, center_estimate, Dfun=self.df, col_deriv=True)[0]

        self.xc, self.yc = center
        ri = self.calc_r(*center)
        self.r = ri.mean()

        return 1 / self.r  # Return the curvature


# color label class
class ColorLabeler:
    def __init__(self):
        # initialize the colors dictionary, containing the color
        # name as the key and the RGB tuple as the value
        colors = OrderedDict({
            "dark skin": (115, 82, 68),
            "light skin": (194, 150, 130),
            "blue sky": (98, 122, 157),
            "foliage": (87, 108, 67),
            "blue flower": (133, 128, 177),
            "bluish green": (103, 189, 170),
            "orange": (214, 126, 44),
            "purplish blue": (8, 91, 166),
            "moderate red": (193, 90, 99),
            "purple": (94, 60, 108),
            "yellow green": (157, 188, 64),
            "orange yellow": (224, 163, 46),
            "blue": (56, 61, 150),
            "green": (70, 148, 73),
            "red": (175, 54, 60),
            "yellow": (231, 199, 31),
            "magneta": (187, 86, 149),
            "cyan": (8, 133, 161),
            "white": (243, 243, 242),
            "neutral 8": (200, 200, 200),
            "neutral 6.5": (160, 160, 160),
            "neutral 5": (122, 122, 121),
            "neutral 3.5": (85, 85, 85),
            "black": (52, 52, 52)})
        # allocate memory for the L*a*b* image, then initialize
        # the color names list
        self.lab = np.zeros((len(colors), 1, 3), dtype="uint8")
        self.colorNames = []
        # loop over the colors dictionary
        for (i, (name, rgb)) in enumerate(colors.items()):
            # update the L*a*b* array and the color names list
            self.lab[i] = rgb
            self.colorNames.append(name)
        # convert the L*a*b* array from the RGB color space
        # to L*a*b*
        self.lab = cv2.cvtColor(self.lab, cv2.COLOR_RGB2LAB)
        #print("color_checker values:{}\n".format(self.lab))

    def label(self, image, c):
            # construct a mask for the contour, then compute the
            # average L*a*b* value for the masked region
            mask = np.zeros(image.shape[:2], dtype="uint8")
            cv2.drawContours(mask, [c], -1, 255, -1)
            mask = cv2.erode(mask, None, iterations=2)
            mean = cv2.mean(image, mask=mask)[:3]

            # initialize the minimum distance found thus far
            minDist = (np.inf, None)
            # loop over the known L*a*b* color values
            for (i, row) in enumerate(self.lab):
                # compute the distance between the current L*a*b*
                # color value and the mean of the image
                d = dist.euclidean(row[0], mean)
                
                #print("mean = {0}, row = {1}, d = {2}, i = {3}\n".format(mean, row[0], d, i)) 
                
                # if the distance is smaller than the current distance,
                # then update the bookkeeping variable
                if d < minDist[0]:
                    minDist = (d, i)
            # return the name of the color with the smallest distance
            return self.colorNames[minDist[1]], mean

    def label_c(self, lab_color_value):
            # initialize the minimum distance found thus far
            minDist = (np.inf, None)
           
            # loop over the known L*a*b* color values
            for (i, row) in enumerate(self.lab):
                # compute the distance between the current L*a*b*
                # color value and the mean of the image
                d = dist.euclidean(row[0], lab_color_value)
                
                #print("mean = {0}, row = {1}, d = {2}, i = {3}\n".format(mean, row[0], d, i)) 
                
                # if the distance is smaller than the current distance,
                # then update the bookkeeping variable
                if d < minDist[0]:
                    minDist = (d, i)
            # return the name of the color with the smallest distance
            return self.colorNames[minDist[1]]


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #printpath + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        
        print("{} path exists!\n".format(path))
        return False
        


# find the closest point wihch minimize the distance between current point and the center of image
def closest_node(pt, pts):
    
    closest_index = dist.cdist([pt], pts).argmin()
    
    return closest_index



# gets the bounding boxes of contours and calculates the distance between two rectangles
def calculate_contour_distance(contour1, contour2): 
    
    x1, y1, w1, h1 = cv2.boundingRect(contour1)
    c_x1 = x1 + w1/2
    c_y1 = y1 + h1/2

    x2, y2, w2, h2 = cv2.boundingRect(contour2)
    c_x2 = x2 + w2/2
    c_y2 = y2 + h2/2

    return max(abs(c_x1 - c_x2) - (w1 + w2)/2, abs(c_y1 - c_y2) - (h1 + h2)/2)


# using numpy.concatenate because each contour is just a numpy array of points
def merge_contours(contour1, contour2):
    
    return np.concatenate((contour1, contour2), axis=0)
    
    #return np.vstack([contour1, contour2])


#group contours such that one contour corresponds to one object.
#when some contours that belong to the same object are detected separately
def agglomerative_cluster(contours, threshold_distance=40.0):
    
    current_contours = contours
    while len(current_contours) > 1:
        min_distance = None
        min_coordinate = None

        for x in range(len(current_contours)-1):
            for y in range(x+1, len(current_contours)):
                distance = calculate_contour_distance(current_contours[x], current_contours[y])
                if min_distance is None:
                    min_distance = distance
                    min_coordinate = (x, y)
                elif distance < min_distance:
                    min_distance = distance
                    min_coordinate = (x, y)

        if min_distance < threshold_distance:
            index1, index2 = min_coordinate
            current_contours[index1] = merge_contours(current_contours[index1], current_contours[index2])
            del current_contours[index2]
        else: 
            break

    return current_contours



# segment foreground object using color clustering method
def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    
    
    
    image_LAB = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
    
    cl = ColorLabeler()
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (height, width, n_channel) = image.shape
    
    if n_channel > 1:
        
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
        
 
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    
    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    # define number of cluster, at lease 2 cluster including background
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    '''
    if args['out_boundary']:
        thresh_cleaned = (thresh)
    
    else:
        
        if np.count_nonzero(thresh) > 0:
            
            thresh_cleaned = clear_border(thresh)
        else:
            thresh_cleaned = thresh
    '''
    
    if np.count_nonzero(thresh) > 0:
        
        #thresh_cleaned = clear_border(thresh)
        thresh_cleaned = thresh
        
    else:
        thresh_cleaned = thresh

    (numLabels, labels, stats, centroids) = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)

    '''
    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    
    # extract the connected component statistics for the current label
    sizes = stats[1:, cv2.CC_STAT_AREA]
    Coord_left = stats[1:, cv2.CC_STAT_LEFT]
    Coord_top = stats[1:, cv2.CC_STAT_TOP]
    Coord_width = stats[1:, cv2.CC_STAT_WIDTH]
    Coord_height = stats[1:, cv2.CC_STAT_HEIGHT]
    Coord_centroids = np.delete(centroids,(0), axis=0)
    

    
    #print("Coord_centroids {}\n".format(centroids[1][1]))
    
    #print("[width, height] {} {}\n".format(width, height))
    
    numLabels = numLabels - 1
    '''

    
    
    ################################################################################################
    

    if args['max_size'] == 1000000:
        
        max_size = width*height
    else:
        max_size = args['max_size']
    
    # initialize an output mask 
    mask = np.zeros(gray.shape, dtype="uint8")
    
    # loop over the number of unique connected component labels, skipping
    # over the first label (as label zero is the background)
    for i in range(1, numLabels):
    # extract the connected component statistics for the current label
        x = stats[i, cv2.CC_STAT_LEFT]
        y = stats[i, cv2.CC_STAT_TOP]
        w = stats[i, cv2.CC_STAT_WIDTH]
        h = stats[i, cv2.CC_STAT_HEIGHT]
        area = stats[i, cv2.CC_STAT_AREA]
    
        
        # ensure the width, height, and area are all neither too small
        # nor too big
        keepWidth = w > 500 and w < 50000
        keepHeight = h > 500 and h < 50000
        keepArea = area > min_size and area < max_size
        
        #if all((keepWidth, keepHeight, keepArea)):
        # ensure the connected component we are examining passes all three tests
        #if all((keepWidth, keepHeight, keepArea)):
        if keepArea:
        # construct a mask for the current connected component and
        # then take the bitwise OR with the mask
            print("[INFO] keeping connected component '{}'".format(i))
            componentMask = (labels == i).astype("uint8") * 255
            mask = cv2.bitwise_or(mask, componentMask)
            

    img_thresh = mask
    
    
    ###################################################################################################
    size_kernel = 10
    
    #if mask contains mutiple non-connected parts, combine them into one. 
    (contours, hier) = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    if len(contours) > 1:
        
        print("mask contains mutiple non-conected parts, combine them into one\n")
        
        kernel = np.ones((size_kernel,size_kernel), np.uint8)

        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        img_thresh = closing
        
        

        
        ###########################################################################
    '''
    if args["cue_color"] == 1:
    
        img_mask = np.zeros([height, width], dtype="uint8")
        
        #img_mask = np.zeros(gray.shape, dtype="uint8")
         
        # filter contours by color cue
        for idx, c in enumerate(contours):
            
            # compute the center of the contour
            M = cv2.moments(c)
            cX = int(M["m10"] / M["m00"])
            cY = int(M["m01"] / M["m00"])
        
            (color_name, color_value) = cl.label(image_LAB, c)
            
            #img_thresh = cv2.putText(img_thresh, "{}".format(color_name), (int(cX), int(cY)), cv2.FONT_HERSHEY_SIMPLEX, 1.8, (255, 0, 0), 2)
            
            print(color_name)
            
            keepColor = color_name == "foliage"  or color_name == "green" 
            
            #or color_name == "dark skin" or color_name == "light skin"

            if keepColor:
                
                #img_mask = cv2.drawContours(img_mask, c, -1, (255), -1)
                img_mask = cv2.drawContours(image=img_mask, contours=[c], contourIdx=-1, color=(255,255,255), thickness=cv2.FILLED)
                #img_mask = cv2.fillPoly(img_mask, pts = [contours], color =(255,255,255))
       
       
            img_thresh = img_mask
    
    '''
        
    
    
    ###################################################################################################
    # use location based selection of plant object, keep the closest componnent  to the center
    '''
    if args["cue_loc"] == 1:
        
        (contours, hier) = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
        if len(contours) > 1:
            
            # location based selection of plant object
            (numLabels, labels, stats, centroids) = cv2.connectedComponentsWithStats(img_thresh, connectivity = 8)

            #keep the center component 

            x_center = int(width // 2)
            y_center = int(height // 2)
            
            Coord_centroids = np.delete(centroids,(0), axis=0)
            
            
            #print("x_center, y_center = {} {}".format(x_center,y_center))
            #print("centroids = {}".format(centroids))
            
            #finding closest point among the grid points list ot the M coordinates
            idx_closest = closest_node((x_center,y_center), Coord_centroids) + 1
            
            print("idx_closest = {}  {}".format(idx_closest, Coord_centroids[idx_closest]))
            
            
            for i in range(1, numLabels):
                
                (cX, cY) = (centroids[i][0], centroids[i][1])
                
                #print(cX, cY)
                
                img_thresh = cv2.putText(img_thresh, "#{}".format(i), (int(cX), int(cY)), cv2.FONT_HERSHEY_SIMPLEX, 1.8, (255, 0, 0), 2)
                
                img_thresh = cv2.putText(img_thresh, "center", (int(x_center), int(y_center)), cv2.FONT_HERSHEY_SIMPLEX, 2.8, (255, 0, 0), 2)
            
            
            
            if numLabels > 1:
                img_thresh = np.zeros([height, width], dtype=np.uint8)
             
                img_thresh[labels == idx_closest] = 255
         
    
    '''
    ###################################################################################################
    #check adjacent contours when mutiple disconnected objects are detected
    

    #return img_thresh    
    #return thresh_cleaned
    
    return img_thresh

    

# compute medial axis from the mask of image
def medial_axis_image(thresh):
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))
    
    image_medial_axis = medial_axis(image_bw)
    
    return image_medial_axis


# compute the skeleton from the mask of image
def skeleton_bw(thresh):

    # Convert mask to boolean image, rather than 0 and 255 for skimage to use it
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))

    #skeleton = morphology.skeletonize(image_bw)
    
    skeleton = morphology.thin(image_bw)
    
    skeleton_img = skeleton.astype(np.uint8) * 255



    return skeleton_img, skeleton


# segmentation using wateshed method
def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = thresh)
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels


# compute percentage as two decimals value
def percentage(part, whole):
  
  #percentage = "{:.0%}".format(float(part)/float(whole))
  
  percentage = "{:.2f}".format(float(part)/float(whole))
  
  return str(percentage)


# convert image from RGB to LAB color space
def image_BRG2LAB(image_file):

   # extarct path and name of the image file
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    
    # extract the base name 
    base_name = os.path.splitext(os.path.basename(filename))[0]

    # get the image file name
    image_file_name = Path(image_file).name
    
    print("Converting image {0} from RGB to LAB color space\n".format(str(image_file_name)))
    
    
    # load the input image 
    image = cv2.imread(image_file)
    
    # change to RGB space
    image_RGB = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    #plt.imshow(image_RGB)
    #plt.show()
    
    # get pixel color
    pixel_colors = image_RGB.reshape((np.shape(image_RGB)[0]*np.shape(image_RGB)[1], 3))
    
    norm = colors.Normalize(vmin=-1.,vmax=1.)
    
    norm.autoscale(pixel_colors)
    
    pixel_colors = norm(pixel_colors).tolist()
    
    #pixel_colors_array = np.asarray(pixel_colors)
    
    #pixel_colors = pixel_colors.ravel()
    
    # change to lab space
    image_LAB = cv2.cvtColor(image, cv2.COLOR_BGR2LAB )
    
   
    (L_chanel, A_chanel, B_chanel) = cv2.split(image_LAB)
    

    ######################################################################
   
    
    fig = plt.figure(figsize=(8.0, 6.0))
    
    axis = fig.add_subplot(1, 1, 1, projection="3d")
    
    axis.scatter(L_chanel.flatten(), A_chanel.flatten(), B_chanel.flatten(), facecolors = pixel_colors, marker = ".")
    axis.set_xlabel("L:ightness")
    axis.set_ylabel("A:red/green coordinate")
    axis.set_zlabel("B:yellow/blue coordinate")
    
    # save segmentation result
    result_file = (result_path + base_name + '_lab' + file_extension)
    
    plt.savefig(result_file, bbox_inches = 'tight', dpi = 1000)
    
    
# detect the circle marker in image
def circle_detection(image):

    """Detecting Circles in Images using OpenCV and Hough Circles
    
    Inputs: 
    
        image: image loaded 

    Returns:
    
        circles: detcted circles
        
        circle_detection_img: circle overlayed with image
        
        diameter_circle: diameter of detected circle
        
    """
    
    # create background image for drawing the detected circle
    output = image.copy()
    
    circle_detection_img = image.copy()
    
    # obtain image dimension
    img_height, img_width, n_channels = image.shape
    
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    #backup input image
    circle_detection_img = image.copy()
    
    # change image from RGB to Gray scale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # apply blur filter 
    blurred = cv2.medianBlur(gray, 25)
    
    # setup parameters for circle detection
    
    # This parameter is the inverse ratio of the accumulator resolution to the image resolution default 1.5
    #(see Yuen et al. for more details). Essentially, the larger the dp gets, the smaller the accumulator array gets.
    dp = 1.0
    
    #Minimum distance between the center (x, y) coordinates of detected circles. 
    #If the minDist is too small, multiple circles in the same neighborhood as the original may be (falsely) detected. 
    #If the minDist is too large, then some circles may not be detected at all.
    minDist = 100
    
    #Gradient value used to handle edge detection in the Yuen et al. method.
    #param1 = 30
    
    #accumulator threshold value for the cv2.HOUGH_GRADIENT method. 
    #The smaller the threshold is, the more circles will be detected (including false circles). 
    #The larger the threshold is, the more circles will potentially be returned. 
    #param2 = 30  
    
    #Minimum/Maximum size of the radius (in pixels).
    #minRadius = 80
    #maxRadius = 120 
    
    # detect circles in the image
    #circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, 1.2, minDist, param1=param1, param2=param2, minRadius=minRadius, maxRadius=maxRadius)
    
    # detect circles in the image
    circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, dp, minDist)
    
    # initialize diameter of detected circle
    diameter_circle = 0
    
    
    circle_center_coord = []
    circle_center_radius = []
    idx_closest = 0
    
    
    # At leaset one circle is found
    if circles is not None:
        
        # Get the (x, y, r) as integers, convert the (x, y) coordinates and radius of the circles to integers
        circles = np.round(circles[0, :]).astype("int")
       
        if len(circles) < 2:
           
            print("Only one circle was found!\n")
           
        else:
            
            print("More than one circles were found!\n")
        
            idx_closest = 0
        
            #cv2.circle(output, (x, y), r, (0, 255, 0), 2)
          
        # loop over the circles and the (x, y) coordinates to get radius of the circles
        for (x, y, r) in circles:
            
            coord = (x, y)
            
            circle_center_coord.append(coord)
            circle_center_radius.append(r)

        if idx_closest == 0:

            print("Circle marker with radius = {} was detected!\n".format(circle_center_radius[idx_closest]))
        
            '''
            # draw the circle in the output image, then draw a center
            circle_detection_img = cv2.circle(circle_detection_img, circle_center_coord[idx_closest], circle_center_radius[idx_closest], (0, 255, 0), 4)
            circle_detection_img = cv2.circle(circle_detection_img, circle_center_coord[idx_closest], 5, (0, 128, 255), -1)
            '''
            
            # compute the diameter of coin
            diameter_circle = circle_center_radius[idx_closest]*2
            
        
            # mask the detected circle with black color
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            tmp_mask = np.zeros((gray.shape), np.uint8)
            
            #tmp_mask = np.zeros([img_width, img_height], dtype=np.uint8)

            tmp_mask = cv2.circle(tmp_mask, circle_center_coord[idx_closest], circle_center_radius[idx_closest] + 50, (255, 255, 255), -1)

            tmp_mask_binary = cv2.threshold(tmp_mask, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
            
            tmp_mask_binary = cv2.bitwise_not(tmp_mask_binary) 
      
            masked_tmp = cv2.bitwise_and(image.copy(), image.copy(), mask = tmp_mask_binary)
            
            #####################################################
            # save marker part as detection results
            (startX, startY) = circle_center_coord[idx_closest]
            
            sx = startX -r*1 if startX -r*1 > 0 else 0
            sy = startY -r*1 if startY -r*1 > 0 else 0
            
            endX = startX + int(r*1.2) 
            endY = startY + int(r*1.2) 
            
            circle_detection_img = output[sy:endY, sx:endX]
            
            
            ###################################################
            
            # crop ROI region based on the location of circle marker
            offset = 1250
            
            endX = startX + int(r*1.2) + offset
            endY = startY + int(r*1.2) + offset
            
            sx = startX -r*4 if startX -r*4 > 0 else 0
            sy = startY -r*4 if startY -r*4 > 0 else 0

            ROI_region = masked_tmp[sy:endY, sx:endX]
        
        #sticker_crop_img = output
    
    else:
        
        print("No circle was found!\n")
        
        ROI_region = output
        
        masked_tmp = output
        
        diameter_circle = 0
    
    return diameter_circle, ROI_region, circle_detection_img


'''
# Detect stickers in the image
def sticker_detect(img_ori):
    

   

    # load the image, clone it for output, and then convert it to grayscale
    img_rgb = img_ori.copy()
    
    # Convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # Store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    
    #(minVal, maxVal, minLoc, maxLoc) = cv2.minMaxLoc(res)
    
    
    # Specify a threshold 
    threshold = 0.6
    
    if np.amax(res) > threshold:
        
        flag = True
    else:

        flag = False
    
    print(flag)
    

    # Store the coordinates of matched area in a numpy array 
    loc = np.where( res >= threshold)  
    
    if len(loc):
    
        (y,x) = np.unravel_index(res.argmax(), res.shape)
    
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)
    
        #print(y,x)
        
        #print(min_val, max_val, min_loc, max_loc)
        
        
        (startX, startY) = max_loc
        endX = startX + template.shape[0] + 1050 + 110
        endY = startY + template.shape[1] + 1050 + 110
        
        
        # Draw a rectangle around the matched region. 
        for pt in zip(*loc[::-1]): 
            
            sticker_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,255,255), 2)
        
        
        sticker_crop_img = img_rgb[startY:endY, startX:endX]


    return  sticker_crop_img, sticker_overlay
'''




# compute the size and shape info of the foreground
def comp_external_contour(orig, thresh):
    
    #find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    thresh_merged = thresh
    
    if len(contours) > 1:
        
        #####################################################################################
        print("Merging contours...\n")
        
        # using agglomerative clustering algorithm to group contours belonging to same object
        contours_list = [element for element in contours]

        # merge adjacent contours with distance threshold
        gp_contours = agglomerative_cluster(contours_list, threshold_distance = 50.0)

        #test_mask = np.zeros([height, width], dtype="uint8")

        thresh_merged = cv2.drawContours(thresh_merged, gp_contours, -1,255,-1)

        #define result path for labeled images
        #result_img_path = file_path + 'test_mask.png'
        #cv2.imwrite(result_img_path, test_mask)
        
        ################################################################################
        #find contours and get the external one
        contours, hier = cv2.findContours(thresh_merged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    else:
        
        #####################################################################################
        print("No need to merge contours...\n")
    
    
    
    # sort the contours based on area from largest to smallest
    contours_sorted = sorted(contours, key = cv2.contourArea, reverse = True)
    
    # initialize parameters
    area_c_cmax = 0
   
    img_height, img_width, img_channels = orig.shape
   
    index = 1
    
    trait_img = orig.copy()
    
    area = 0
    
    compactness = 0
    
    solidity = 0
    
    w=h=0
    
    
    for index, c in enumerate(contours_sorted):
        
    #for c in contours:
        if index < 1:
    
            #get the bounding rect
            x, y, w, h = cv2.boundingRect(c)
            
            if w>0 and h>0:
                
                # draw contour
                trait_img = cv2.drawContours(orig, contours, -1, (0, 255, 255), 1)
        
                # draw a green rectangle to visualize the bounding rect
                roi = orig[y:y+h, x:x+w]
                
                print("ROI {} detected ...\n".format(index))
                #result_file = (result_path +  str(index) + file_extension)
                #cv2.imwrite(result_file, roi)
                
                #trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 3)
                
                index+= 1

                
                #get the min area rect
                rect = cv2.minAreaRect(c)
                
                (x, y), (minAreaRect_width, minAreaRect_height), angle = rect
                
                box = cv2.boxPoints(rect)
                # convert all coordinates floating point values to int
                box = np.int0(box)
                #draw a red 'nghien' rectangle
                trait_img = cv2.drawContours(orig, [box], 0, (255, 255, 0), 1)
                
                
                
                 # get convex hull
                hull = cv2.convexHull(c)
                # draw it in red color
                trait_img = cv2.drawContours(orig, [hull], -1, (0, 255, 0), 1)
                
                '''
                # calculate epsilon base on contour's perimeter
                # contour's perimeter is returned by cv2.arcLength
                epsilon = 0.01 * cv2.arcLength(c, True)
                # get approx polygons
                approx = cv2.approxPolyDP(c, epsilon, True)
                # draw approx polygons
                trait_img = cv2.drawContours(orig, [approx], -1, (0, 255, 0), 1)
             
                # hull is convex shape as a polygon
                hull = cv2.convexHull(c)
                trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255))
                '''
                
                '''
                #get the min enclosing circle
                (x, y), radius = cv2.minEnclosingCircle(c)
                # convert all values to int
                center = (int(x), int(y))
                radius = int(radius)
                # and draw the circle in blue
                trait_img = cv2.circle(orig, center, radius, (255, 0, 0), 2)
                '''
                
                area = cv2.contourArea(c)
                print("Leaf area = {0:.2f}... \n".format(area))
                
                
                hull = cv2.convexHull(c)
                hull_area = cv2.contourArea(hull)
                
                compactness = float(area)/hull_area
                print("compactness = {0:.2f}... \n".format(compactness))
                
                
                solidity = float(area)/(w*h)
                print("solidity = {0:.2f}... \n".format(solidity))
                
                extLeft = tuple(c[c[:,:,0].argmin()][0])
                extRight = tuple(c[c[:,:,0].argmax()][0])
                extTop = tuple(c[c[:,:,1].argmin()][0])
                extBot = tuple(c[c[:,:,1].argmax()][0])
                
                trait_img = cv2.circle(orig, extLeft, 6, (255, 0, 0), -1)
                trait_img = cv2.circle(orig, extRight, 6, (255, 0, 0), -1)
                trait_img = cv2.circle(orig, extTop, 6, (255, 0, 0), -1)
                trait_img = cv2.circle(orig, extBot, 6, (255, 0, 0), -1)
                
                max_width = dist.euclidean(extLeft, extRight)
                max_height = dist.euclidean(extTop, extBot)
                
                if max_width > max_height:
                    trait_img = cv2.line(orig, extLeft, extRight, (0,0,255), 1)
                else:
                    trait_img = cv2.line(orig, extTop, extBot, (0,0,255), 1)
                
                
                longest_dimension = max(max_width, max_height)

                print("Width and height are {0:.2f},{1:.2f}... \n".format(w, h))
        
        
            
    return trait_img, area, compactness, solidity, minAreaRect_width, minAreaRect_height, longest_dimension
    
    
    
    


# rgb to hex conversion
def RGB2HEX(color):
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))
    
# rgb to float conversion
def RGB2FLOAT(color):
    return "{:.2f}{:.2f}{:.2f}".format(int(color[0]/255.0), int(color[1]/255.0), int(color[2]/255.0))


# get color map from index
def get_cmap(n, name = 'hsv'):
    '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    RGB color; the keyword argument name must be a standard mpl colormap name.'''
    return plt.get_cmap(name, n)
    
    #import matplotlib.cm
    #return matplotlib.cm.get_cmap(name, n)
    




# normalzie image
def _normalise_image(image, *, image_cmap=None):
    image = img_as_float(image)
    if image.ndim == 2:
        if image_cmap is None:
            image = gray2rgb(image)
        else:
            image = plt.get_cmap(image_cmap)(image)[..., :3]
    return image



# compute the image brightness
def isbright(image_file):
    
    # Set up threshold value for luminous channel, can be adjusted and generalized 
    thresh = 1.5
    
    # Load image file 
    orig = cv2.imread(image_file)
    
    # Make backup image
    image = orig.copy()
    
    # Get file name
    #abs_path = os.path.abspath(image_file)
    
    #filename, file_extension = os.path.splitext(abs_path)
    #base_name = os.path.splitext(os.path.basename(filename))[0]

    image_file_name = Path(image_file).name
    
    # Convert color space to LAB format and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
    
    # Normalize L channel by dividing all pixel values with maximum pixel value
    L = L/np.max(L)
    
    text_bool = "bright" if np.mean(L) < thresh else "dark"
    
    #return image_file_name, np.mean(L), text_bool
    
    #print("np.mean(L) < thresh = {}".format(np.mean(L)))
    
    #return np.mean(L) < thresh
    
    b_bright = 1.0 < thresh
    
    #print("Image brightness: {}\n".foramt(b_bright))
    
    return  b_bright, np.mean(L)
    

# compute the image brightness
def remove_character_string(str_input):
    
    return str_input.replace('#', '')


# compute the mean value of hex colors
def hex_mean_color(color_list):
    
    average_value = (int(remove_character_string(color1), 16) + int(remove_character_string(color2), 16) + int(remove_character_string(color3), 16) + int(remove_character_string(color4), 16))//4
       
    return hex(average_value)




# convert from RGB to LAB space,
# Convert it to LAB color space to access the luminous channel which is independent of colors.
def RGB2LAB(image, mask):
    
    # Make backup image
    image_rgb = image.copy()
    
    # apply object mask
    masked_rgb = cv2.bitwise_and(image_rgb, image_rgb, mask = mask)
    
    # Convert color space to LAB space and extract L channel
    (L, A, B) = cv2.split(cv2.cvtColor(masked_rgb, cv2.COLOR_BGR2LAB))
    

    return masked_rgb, L, A, B
    


#computation of color_difference index
def color_diff_index(ref_color, rgb_colors):
    
    #print(ref_color)
    
    #lab_colors = cv2.cvtColor(rgb_colors, cv2.COLOR_RGB2LAB)
    
    color_diff = []
    
    for index, value in enumerate(rgb_colors): 
        
        curr_color = rgb2lab(value)
        
        # color value from skimage rgb2lab: ranges of Lab values which are: L (0-100), a (-128-127), b (-128-127). 
        # differnt from OpenCV cv2.COLOR_RGB2LAB, need scale to 0~255
        curr_color_scaled = (curr_color + [155, 128, 128]) 

        #color difference in CIE lab space
        diff = deltaE_cie76(ref_color, curr_color_scaled)
        
        diff_value = float(diff)

        #diff = dist.euclidean(std_color_value[1].flatten(), checker_color_value[13].flatten())
        
        color_diff.append(diff_value)
        
        #print("current color value = {}, cluster index = {}, color difference = {}: \n".format(curr_color_scaled, index, diff)) 
    
    #color_diff_index = sum(color_diff) / len(color_diff)
    
        
    return color_diff


# Max RGB filter 
def max_rgb_filter(image):
    
    # split the image into its BGR components
    (B, G, R) = cv2.split(image)
    
    # find the maximum pixel intensity values for each
    # (x, y)-coordinate,, then set all pixel values less
    # than M to zero
    M = np.maximum(np.maximum(R, G), B)
    R[R < M] = 0
    G[G < M] = 0
    B[B < M] = 0
    
    
    # merge the channels back together and return the image
    return cv2.merge([B, G, R])


    

# get file information from the file path using python3
def get_file_info(file_full_path):
    
    p = pathlib.Path(file_full_path)

    filename = p.name

    basename = p.stem

    file_path = p.parent.absolute()

    file_path = os.path.join(file_path, '')
    
    return file_path, filename, basename



# execute script inside program
def execute_script(cmd_line):
    
    try:
        #print(cmd_line)
        #os.system(cmd_line)

        process = subprocess.getoutput(cmd_line)
        
        print(process)
        
        #process = subprocess.Popen(cmd_line, shell = True, stdout = subprocess.PIPE)
        #process.wait()
        #print(process.communicate())
        
    except OSError:
        
        print("Failed ...!\n")



# sort contours based on user defined method
def sort_contours(cnts, method = "left-to-right"):
    
    """sort contours based on user defined method
    
    Inputs: 
    
        cnts: contours extracted from mask image
        
        method: user defined method, default was "left-to-right"
        

    Returns:
    
        sorted_cnts: list of sorted contours 
        
    """   
    
    
    # initialize the reverse flag and sort index
    reverse = False
    i = 0
    
    # handle if we need to sort in reverse
    if method == "right-to-left" or method == "bottom-to-top":
        reverse = True
        
    # handle if we are sorting against the y-coordinate rather than
    # the x-coordinate of the bounding box
    if method == "top-to-bottom" or method == "bottom-to-top":
        i = 1
        
    # construct the list of bounding boxes and sort them from top to bottom
    boundingBoxes = [cv2.boundingRect(c) for c in cnts]
    
    (sorted_cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes), key=lambda b:b[1][i], reverse=reverse))
    
    # return the list of sorted contours and bounding boxes
    return sorted_cnts






# find contours in binary image mask and sort them in left-to-right order
def get_contours(image_thresh):
    
    """find contours in binary image mask and sort them in left-to-right order
    
    Inputs: 
    
        image_thresh: image mask

    Returns:
    
        cnts_sorted:  sorted contours
        
    """
    # find contours in the thresholded image
    cnts = cv2.findContours(image_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    cnts = imutils.grab_contours(cnts)

    # sort the contour based on area size from largest to smallest, and get the first two max contours
    cnts_sorted = sorted(cnts, key = cv2.contourArea, reverse = True)[0:n_ear]

    # sort the contours from left to right
    cnts_sorted = sort_contours(cnts_sorted, method = "left-to-right")

   
    print("Sorting {} objects in left-to-right order\n".format(len(cnts_sorted)))
    
    return cnts_sorted



# compute all the traits
def extract_traits(image_file, result_path):


    #gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    
    base_name = os.path.splitext(os.path.basename(filename))[0]

    file_size = int(os.path.getsize(image_file)/MBFACTOR)
    
    image_file_name = Path(image_file).name
   
    
    # make the folder to store the results
    #current_path = abs_path + '/'
    
    print("Computing traits for image : {0}\n".format(str(image_file_name)))
    

    
    #########################################################################################
    # check color brightness
    #(b_bright, b_value) = isbright(image_file)
    
    # initilize parameters
    area = solidity = max_width = max_height = avg_curv = n_leaves = diameter_circle = compactness = longest_dimension = 0
        

    ################################################################################
    # load image data
    image = cv2.imread(image_file)
    
    # make backup image
    orig = image.copy()
    
    # get the dimension of the image
    img_height, img_width, img_channels = orig.shape

    #source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)

    #QR_data = 0
    
    ################################################################################
    # output image file info
    if image is not None:
        
        print("Plant object segmentation using automatic color clustering method... \n")
        
        print("Image file size: {} MB, dimension: {} X {}, channels : {}\n".format(str(file_size), img_height, img_width, img_channels))
    
    
   
        
    ##########################################################################
    #Plant region detection (defined as ROI_region)

    #roi_image = ROI_region.copy()
    
    ROI_region = orig.copy()
    

        
    ###########################################################################################################
    #compute external contour, shape info
    
    print(ROI_region.shape)
    
    #thresh = cv2.threshold(ROI_region, 0, 255, cv2.THRESH_BINARY)

    
    # Convert the image to grayscale
    gray = cv2.cvtColor(ROI_region, cv2.COLOR_BGR2GRAY)
    
    ret, binary_mask = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY)
    
    if cv2.countNonZero(binary_mask) == 0:
    
        print("Image is black\n")
        
        area = solidity = max_width = max_height = compactness = longest_dimension = 0
        
        trait_img = image

    else:
        
        (trait_img, area, compactness, solidity, max_width, max_height, longest_dimension) = comp_external_contour(ROI_region, binary_mask)
        '''

        contours, _ = cv2.findContours(binary_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # Assuming the largest contour represents the object of interest
        if contours:
            largest_contour = max(contours, key=cv2.contourArea)
            area_contour = cv2.contourArea(largest_contour)
        else:
            area_contour = 0 # No contours found
            
        print(area_contour)
        '''
        
        
        #result_img_path = result_path + 'binary_mask.png'
        #cv2.imwrite(result_img_path, binary_mask)
        
        
        
        
    
    return image_file_name, area, solidity, max_width, max_height, compactness, longest_dimension, trait_img
    


# save result files
def write_image_output(imagearray, result_path, base_name, addition, ext):

    # save segmentation result
    result_file = (result_path + base_name + addition + ext)
    
    #print(result_file)
    
    cv2.imwrite(result_file, imagearray)
    
    # check saved file
    if os.path.exists(result_file):
        print("Result file was saved at {0}\n".format(result_file))

    else:
        print("Result file writing failed!\n")
    




# save results as excel file
def write_excel_output(trait_file, result_list):
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
        #sheet_leaf = wb.create_sheet()
        
        #sheet_leaf.delete_rows(2, sheet_leaf.max_row+1) # for entire sheet


    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet_leaf = wb.create_sheet()

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'leaf_area'
        sheet.cell(row = 1, column = 3).value = 'solidity'
        sheet.cell(row = 1, column = 4).value = 'max_width'
        sheet.cell(row = 1, column = 5).value = 'max_height'
        sheet.cell(row = 1, column = 6).value = 'compactness'
        sheet.cell(row = 1, column = 7).value = 'longest_dimension'
        
        #return image_file_name, area, solidity, max_width, max_height, compactness, longest_dimension, trait_img
        
    for row in result_list:
        sheet.append(row)
   

    #save the csv file
    wb.save(trait_file)





if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", dest = "path", type = str, required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", dest = "filetype", type = str, required = False, default='jpg,png', help = "Image file type")
    ap.add_argument("-o", "--output_path", dest = "output_path", type = str, required = False,    help = "result path")
    ap.add_argument('-s', '--color_space', dest = "color_space", type = str, required = False, default ='lab', help='Color space to use: BGR, HSV, Lab, YCrCb (YCC)')
    ap.add_argument('-c', '--channels', dest = "channels", type = str, required = False, default='2', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num_clusters', dest = "num_clusters", type = int, required = False, default = 4,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    ap.add_argument('-min', '--min_size', dest = "min_size", type = int, required = False, default = 100,  help = 'min size of object to be segmented.')
    ap.add_argument('-max', '--max_size', dest = "max_size", type = int, required = False, default = 1000000,  help = 'max size of object to be segmented.')
    ap.add_argument('-md', '--min_dist', dest = "min_dist", type = int, required = False, default = 15,  help = 'distance threshold of watershed segmentation.')
    ap.add_argument("-da", "--diagonal", dest = "diagonal", type = float, required = False,  default = math.sqrt(2), help = "diagonal line length(cm) of indiviudal color checker module")
    ap.add_argument("-d", '--debug', dest = 'debug', type = int, required = False,  default = 1, help = "Whehter save image results or not, 1 = yes, 0 = no")
    #ap.add_argument("-cc", "--cue_color", dest = "cue_color", type = int, required = False,  default = 0, help="use color cue to detect plant object")
    #ap.add_argument("-cl", "--cue_loc", dest = "cue_loc", type = int, required = False,  default = 0, help="use location cue to detect plant object")
    #ap.add_argument("-ob", "--out_boundary", dest = "out_boundary", type = int, required = False,  default = 0, help="whether the plant object was out of the image boudary or not, 1 yes, 0 no, default 0")
    
    args = vars(ap.parse_args())
    

    # input path
    file_path = args["path"]
    
    ext = args['filetype'].split(',') if 'filetype' in args else []
    
    patterns = [os.path.join(file_path, f"*.{p}") for p in ext]
    
    files = [f for fs in [glob.glob(pattern) for pattern in patterns] for f in fs]
    

    if len(files) > 0:
        
        # check image file format 
        extension_type = check_file_type(file_path, None)
        
        print("Number of input images: {}, Image format: {}\n".format(len(files), extension_type))
    
    else:
        
        print("Input folder was empty...\n")
        
        sys.exit(1)
        

    #print("Input image format: {}\n".format(extension_type))
    
    #print("Input images: {}\n".format(files))
    
    
    # result path
    result_path = args["output_path"] if args["output_path"] is not None else file_path
    
    result_path = os.path.join(result_path, '')

    # printout result path
    print("Output path: {}\n".format(result_path))
    
    
    ########################################################################
    #parameters
    
    min_size = args['min_size']

    min_distance_value = args['min_dist']
    
    diagonal_line_length = args['min_dist']
    
    num_clusters = args['num_clusters'] 
    
    
    args_colorspace = args['color_space']
    args_channels = args['channels']
    args_num_clusters = args['num_clusters']
    

    #accquire image file list
    #image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(files)
    
    

    #print((imgList))
    #global result_path
    
    n_images = len(imgList)
    
    result_list = []
    

    
    
    # loop execute
    for image_id, image in enumerate(imgList):
        
        # get file information
        (file_path, filename, basename) = get_file_info(image)
        
        print("Processing image {} ... \n".format(file_path))
        
        # main pipeline
        (filename, area, solidity, max_width, max_height, compactness, longest_dimension, trait_img) = extract_traits(image, file_path)
        
        result_list.append([filename, area, solidity, max_width, max_height, compactness, longest_dimension])
        
        '''
        if args["debug"] == 1:

            
            file_extension = '.png'
            
            mkpath = os.path.dirname(result_path) +'/' + basename

            mkdir(mkpath)

            image_save_path = mkpath + '/'
            
            

            print("image_save_path: {}\n".format(image_save_path))

            # save segmentation result
            write_image_output(trait_img, image_save_path, basename, '_trait_img', file_extension)
            
            
        '''


    #########################################################################
    #trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    #output in command window in a sum table
 
    #table = tabulate(result_list, headers = ['filename', 'area', 'solidity', 'max_width', 'max_height' ,'avg_curv', 'n_leaves', 'cluster 1', 'cluster 2', 'cluster 3', 'cluster 4', 'cluster 1 hex value', 'cluster 2 hex value', 'cluster 3 hex value', 'cluster 4 hex value'], tablefmt = 'orgtbl')

    #print(table + "\n")
    

    #file_path_full = os.path.join(file_path, '')
    
    

    
    ########################################################################################
    trait_file = (result_path + str(pathlib.PurePath(result_path).name) + '_trait.xlsx')
    #trait_file = (result_path + basename + '_trait.xlsx')

    write_excel_output(trait_file, result_list)

    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
    else:
        print("Error in saving Result file\n")
    
    '''
    #####################################################################################
    # grants read and write access to all result folders
    print("Make result files accessible...\n")

    access_grant = "chmod 777 -R " + result_path 
    
    print(access_grant + '\n')
    
    execute_script(access_grant)
    '''
    
    

    
