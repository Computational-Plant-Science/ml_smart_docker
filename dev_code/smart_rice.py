'''
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract plant shoot traits (larea, solidity, max_width, max_height, avg_curv, color_cluster) 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2024-11-29

USAGE:

    python3 smart_rice.py -p /input/ -o /output/

'''

# import the necessary packages
import subprocess, os, glob, sys
import utils

from collections import Counter
from collections import OrderedDict

import argparse

from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from skimage.feature import peak_local_max
from skimage.morphology import medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure
from skimage.color import rgb2lab, deltaE_cie76
from skimage import morphology
from skimage.segmentation import clear_border, watershed
from skimage.measure import regionprops

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage
from scipy.interpolate import interp1d

#from skan import skeleton_to_csgraph, Skeleton, summarize, draw

#import networkx as nx

import imutils

import numpy as np
import argparse
import cv2

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import math
import openpyxl
import csv
    
from tabulate import tabulate

import warnings
warnings.filterwarnings("ignore")

import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from pathlib import Path

from rembg import remove

from matplotlib import collections

import matplotlib.colors

import pathlib


MBFACTOR = float(1<<20)


# check file type
def check_file_type(image_folder_path, allowed_extensions=None):
    
    if allowed_extensions is None:
        allowed_extensions = ['.jpg', '.png', '.jpeg']

    no_files_in_folder = len(glob.glob(image_folder_path+"/*")) 
    extension_type = ""
    no_files_allowed = 0

    for ext in allowed_extensions:
      no_files_allowed = len(glob.glob(image_folder_path+"/*"+ext))
      if no_files_allowed > 0:
        extension_type = ext

    return extension_type


# curvature computation calss
class ComputeCurvature:

    def __init__(self,x,y):
        """ Initialize some variables """
        self.xc = 0  # X-coordinate of circle center
        self.yc = 0  # Y-coordinate of circle center
        self.r = 0   # Radius of the circle
        self.xx = np.array([])  # Data points
        self.yy = np.array([])  # Data points
        self.x = x  # X-coordinate of circle center
        self.y = y  # Y-coordinate of circle center

    def calc_r(self, xc, yc):
        """ calculate the distance of each 2D points from the center (xc, yc) """
        return np.sqrt((self.xx-xc)**2 + (self.yy-yc)**2)

    def f(self, c):
        """ calculate the algebraic distance between the data points and the mean circle centered at c=(xc, yc) """
        ri = self.calc_r(*c)
        return ri - ri.mean()

    def df(self, c):
        """ Jacobian of f_2b
        The axis corresponding to derivatives must be coherent with the col_deriv option of leastsq"""
        xc, yc = c
        df_dc = np.empty((len(c), self.x.size))

        ri = self.calc_r(xc, yc)
        df_dc[0] = (xc - self.x)/ri                   # dR/dxc
        df_dc[1] = (yc - self.y)/ri                   # dR/dyc
        df_dc = df_dc - df_dc.mean(axis=1)[:, np.newaxis]
        return df_dc

    def fit(self, xx, yy):
        self.xx = xx
        self.yy = yy
        center_estimate = np.r_[np.mean(xx), np.mean(yy)]
        center = optimize.leastsq(self.f, center_estimate, Dfun=self.df, col_deriv=True)[0]

        self.xc, self.yc = center
        ri = self.calc_r(*center)
        self.r = ri.mean()

        return 1 / self.r  # Return the curvature


# color label class
class ColorLabeler:
    def __init__(self):
        # initialize the colors dictionary, containing the color
        # name as the key and the RGB tuple as the value
        colors = OrderedDict({
            "dark skin": (115, 82, 68),
            "light skin": (194, 150, 130),
            "blue sky": (98, 122, 157),
            "foliage": (87, 108, 67),
            "blue flower": (133, 128, 177),
            "bluish green": (103, 189, 170),
            "orange": (214, 126, 44),
            "purplish blue": (8, 91, 166),
            "moderate red": (193, 90, 99),
            "purple": (94, 60, 108),
            "yellow green": (157, 188, 64),
            "orange yellow": (224, 163, 46),
            "blue": (56, 61, 150),
            "green": (70, 148, 73),
            "red": (175, 54, 60),
            "yellow": (231, 199, 31),
            "magneta": (187, 86, 149),
            "cyan": (8, 133, 161),
            "white": (243, 243, 242),
            "neutral 8": (200, 200, 200),
            "neutral 6.5": (160, 160, 160),
            "neutral 5": (122, 122, 121),
            "neutral 3.5": (85, 85, 85),
            "black": (52, 52, 52)})
        # allocate memory for the L*a*b* image, then initialize
        # the color names list
        self.lab = np.zeros((len(colors), 1, 3), dtype="uint8")
        self.colorNames = []
        # loop over the colors dictionary
        for (i, (name, rgb)) in enumerate(colors.items()):
            # update the L*a*b* array and the color names list
            self.lab[i] = rgb
            self.colorNames.append(name)
        # convert the L*a*b* array from the RGB color space
        # to L*a*b*
        self.lab = cv2.cvtColor(self.lab, cv2.COLOR_RGB2LAB)
        #print("color_checker values:{}\n".format(self.lab))

    def label(self, image, c):
            # construct a mask for the contour, then compute the
            # average L*a*b* value for the masked region
            mask = np.zeros(image.shape[:2], dtype="uint8")
            cv2.drawContours(mask, [c], -1, 255, -1)
            mask = cv2.erode(mask, None, iterations=2)
            mean = cv2.mean(image, mask=mask)[:3]

            # initialize the minimum distance found thus far
            minDist = (np.inf, None)
            # loop over the known L*a*b* color values
            for (i, row) in enumerate(self.lab):
                # compute the distance between the current L*a*b*
                # color value and the mean of the image
                d = dist.euclidean(row[0], mean)
                
                #print("mean = {0}, row = {1}, d = {2}, i = {3}\n".format(mean, row[0], d, i)) 
                
                # if the distance is smaller than the current distance,
                # then update the bookkeeping variable
                if d < minDist[0]:
                    minDist = (d, i)
            # return the name of the color with the smallest distance
            return self.colorNames[minDist[1]], mean

    def label_c(self, lab_color_value):
            # initialize the minimum distance found thus far
            minDist = (np.inf, None)
           
            # loop over the known L*a*b* color values
            for (i, row) in enumerate(self.lab):
                # compute the distance between the current L*a*b*
                # color value and the mean of the image
                d = dist.euclidean(row[0], lab_color_value)
                
                #print("mean = {0}, row = {1}, d = {2}, i = {3}\n".format(mean, row[0], d, i)) 
                
                # if the distance is smaller than the current distance,
                # then update the bookkeeping variable
                if d < minDist[0]:
                    minDist = (d, i)
            # return the name of the color with the smallest distance
            return self.colorNames[minDist[1]]


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #printpath + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        
        print("{} path exists!\n".format(path))
        return False
        


# find the closest point wihch minimize the distance between current point and the center of image
def closest_node(pt, pts):
    
    closest_index = dist.cdist([pt], pts).argmin()
    
    return closest_index



# gets the bounding boxes of contours and calculates the distance between two rectangles
def calculate_contour_distance(contour1, contour2): 
    
    x1, y1, w1, h1 = cv2.boundingRect(contour1)
    c_x1 = x1 + w1/2
    c_y1 = y1 + h1/2

    x2, y2, w2, h2 = cv2.boundingRect(contour2)
    c_x2 = x2 + w2/2
    c_y2 = y2 + h2/2

    return max(abs(c_x1 - c_x2) - (w1 + w2)/2, abs(c_y1 - c_y2) - (h1 + h2)/2)


# using numpy.concatenate because each contour is just a numpy array of points
def merge_contours(contour1, contour2):
    
    return np.concatenate((contour1, contour2), axis=0)
    
    #return np.vstack([contour1, contour2])


#group contours such that one contour corresponds to one object.
#when some contours that belong to the same object are detected separately
def agglomerative_cluster(contours, threshold_distance=40.0):
    
    current_contours = contours
    while len(current_contours) > 1:
        min_distance = None
        min_coordinate = None

        for x in range(len(current_contours)-1):
            for y in range(x+1, len(current_contours)):
                distance = calculate_contour_distance(current_contours[x], current_contours[y])
                if min_distance is None:
                    min_distance = distance
                    min_coordinate = (x, y)
                elif distance < min_distance:
                    min_distance = distance
                    min_coordinate = (x, y)

        if min_distance < threshold_distance:
            index1, index2 = min_coordinate
            current_contours[index1] = merge_contours(current_contours[index1], current_contours[index2])
            del current_contours[index2]
        else: 
            break

    return current_contours



# segment foreground object using color clustering method
def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    
    
    
    image_LAB = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
    
    cl = ColorLabeler()
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (height, width, n_channel) = image.shape
    
    if n_channel > 1:
        
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
        
 
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    
    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    # define number of cluster, at lease 2 cluster including background
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    '''
    if args['out_boundary']:
        thresh_cleaned = (thresh)
    
    else:
        
        if np.count_nonzero(thresh) > 0:
            
            thresh_cleaned = clear_border(thresh)
        else:
            thresh_cleaned = thresh
    '''
    
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned = clear_border(thresh)
        #thresh_cleaned = thresh
        
    else:
        thresh_cleaned = thresh

    (numLabels, labels, stats, centroids) = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)

    '''
    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    
    # extract the connected component statistics for the current label
    sizes = stats[1:, cv2.CC_STAT_AREA]
    Coord_left = stats[1:, cv2.CC_STAT_LEFT]
    Coord_top = stats[1:, cv2.CC_STAT_TOP]
    Coord_width = stats[1:, cv2.CC_STAT_WIDTH]
    Coord_height = stats[1:, cv2.CC_STAT_HEIGHT]
    Coord_centroids = np.delete(centroids,(0), axis=0)
    

    
    #print("Coord_centroids {}\n".format(centroids[1][1]))
    
    #print("[width, height] {} {}\n".format(width, height))
    
    numLabels = numLabels - 1
    '''

    
    
    ################################################################################################
    

    if args['max_size'] == 1000000:
        
        max_size = width*height
    else:
        max_size = args['max_size']
    
    # initialize an output mask 
    mask = np.zeros(gray.shape, dtype="uint8")
    
    # loop over the number of unique connected component labels, skipping
    # over the first label (as label zero is the background)
    for i in range(1, numLabels):
    # extract the connected component statistics for the current label
        x = stats[i, cv2.CC_STAT_LEFT]
        y = stats[i, cv2.CC_STAT_TOP]
        w = stats[i, cv2.CC_STAT_WIDTH]
        h = stats[i, cv2.CC_STAT_HEIGHT]
        area = stats[i, cv2.CC_STAT_AREA]
    
        
        # ensure the width, height, and area are all neither too small
        # nor too big
        keepWidth = w > 100 and w < 50000
        keepHeight = h > 100 and h < 50000
        keepArea = area > min_size and area < max_size
        
        #if all((keepWidth, keepHeight, keepArea)):
        # ensure the connected component we are examining passes all three tests
        #if all((keepWidth, keepHeight, keepArea)):
        if keepArea:
        # construct a mask for the current connected component and
        # then take the bitwise OR with the mask
            print("[INFO] keeping connected component '{}'".format(i))
            componentMask = (labels == i).astype("uint8") * 255
            mask = cv2.bitwise_or(mask, componentMask)
            

    img_thresh = mask
    
    
    ###################################################################################################
    size_kernel = 1
    
    #if mask contains mutiple non-connected parts, combine them into one. 
    (contours, hier) = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    if len(contours) > 1:
        
        print("mask contains mutiple non-conected parts, combine them into one\n")
        
        kernel = np.ones((size_kernel,size_kernel), np.uint8)

        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        img_thresh = closing
        



    #return img_thresh    
    #return thresh_cleaned
    
    return img_thresh

    

# compute medial axis from the mask of image
def medial_axis_image(thresh):
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))
    
    image_medial_axis = medial_axis(image_bw)
    
    return image_medial_axis


# compute the skeleton from the mask of image
def skeleton_bw(thresh):

    # Convert mask to boolean image, rather than 0 and 255 for skimage to use it
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))

    #skeleton = morphology.skeletonize(image_bw)
    
    skeleton = morphology.thin(image_bw)
    
    skeleton_img = skeleton.astype(np.uint8) * 255



    return skeleton_img, skeleton


# segmentation using wateshed method
def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = thresh)
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels


# compute percentage as two decimals value
def percentage(part, whole):
  
  #percentage = "{:.0%}".format(float(part)/float(whole))
  
  percentage = "{:.2f}".format(float(part)/float(whole))
  
  return str(percentage)


# convert image from RGB to LAB color space
def image_BRG2LAB(image_file):

   # extarct path and name of the image file
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    
    # extract the base name 
    base_name = os.path.splitext(os.path.basename(filename))[0]

    # get the image file name
    image_file_name = Path(image_file).name
    
    print("Converting image {0} from RGB to LAB color space\n".format(str(image_file_name)))
    
    
    # load the input image 
    image = cv2.imread(image_file)
    
    # change to RGB space
    image_RGB = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    #plt.imshow(image_RGB)
    #plt.show()
    
    # get pixel color
    pixel_colors = image_RGB.reshape((np.shape(image_RGB)[0]*np.shape(image_RGB)[1], 3))
    
    norm = colors.Normalize(vmin=-1.,vmax=1.)
    
    norm.autoscale(pixel_colors)
    
    pixel_colors = norm(pixel_colors).tolist()
    
    #pixel_colors_array = np.asarray(pixel_colors)
    
    #pixel_colors = pixel_colors.ravel()
    
    # change to lab space
    image_LAB = cv2.cvtColor(image, cv2.COLOR_BGR2LAB )
    
   
    (L_chanel, A_chanel, B_chanel) = cv2.split(image_LAB)
    

    ######################################################################
   
    
    fig = plt.figure(figsize=(8.0, 6.0))
    
    axis = fig.add_subplot(1, 1, 1, projection="3d")
    
    axis.scatter(L_chanel.flatten(), A_chanel.flatten(), B_chanel.flatten(), facecolors = pixel_colors, marker = ".")
    axis.set_xlabel("L:ightness")
    axis.set_ylabel("A:red/green coordinate")
    axis.set_zlabel("B:yellow/blue coordinate")
    
    # save segmentation result
    result_file = (result_path + base_name + '_lab' + file_extension)
    
    plt.savefig(result_file, bbox_inches = 'tight', dpi = 1000)
    
    
# detect the circle marker in image
def circle_detection(image):

    """Detecting Circles in Images using OpenCV and Hough Circles
    
    Inputs: 
    
        image: image loaded 

    Returns:
    
        circles: detcted circles
        
        circle_detection_img: circle overlayed with image
        
        diameter_circle: diameter of detected circle
        
    """
    
    # create background image for drawing the detected circle
    output = image.copy()
    
    circle_detection_img = image.copy()
    
    # obtain image dimension
    img_height, img_width, n_channels = image.shape
    
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    #backup input image
    circle_detection_img = image.copy()
    
    # change image from RGB to Gray scale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # apply blur filter 
    blurred = cv2.medianBlur(gray, 25)
    
    # setup parameters for circle detection
    
    # This parameter is the inverse ratio of the accumulator resolution to the image resolution default 1.5
    #(see Yuen et al. for more details). Essentially, the larger the dp gets, the smaller the accumulator array gets.
    dp = 1.0
    
    #Minimum distance between the center (x, y) coordinates of detected circles. 
    #If the minDist is too small, multiple circles in the same neighborhood as the original may be (falsely) detected. 
    #If the minDist is too large, then some circles may not be detected at all.
    minDist = 100
    
    #Gradient value used to handle edge detection in the Yuen et al. method.
    #param1 = 30
    
    #accumulator threshold value for the cv2.HOUGH_GRADIENT method. 
    #The smaller the threshold is, the more circles will be detected (including false circles). 
    #The larger the threshold is, the more circles will potentially be returned. 
    #param2 = 30  
    
    #Minimum/Maximum size of the radius (in pixels).
    #minRadius = 80
    #maxRadius = 120 
    
    # detect circles in the image
    #circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, 1.2, minDist, param1=param1, param2=param2, minRadius=minRadius, maxRadius=maxRadius)
    
    # detect circles in the image
    circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, dp, minDist)
    
    # initialize diameter of detected circle
    diameter_circle = 0
    
    
    circle_center_coord = []
    circle_center_radius = []
    idx_closest = 0
    
    
    # At leaset one circle is found
    if circles is not None:
        
        # Get the (x, y, r) as integers, convert the (x, y) coordinates and radius of the circles to integers
        circles = np.round(circles[0, :]).astype("int")
       
        if len(circles) < 2:
           
            print("Only one circle was found!\n")
           
        else:
            
            print("More than one circles were found!\n")
        
            idx_closest = 0
        
            #cv2.circle(output, (x, y), r, (0, 255, 0), 2)
          
        # loop over the circles and the (x, y) coordinates to get radius of the circles
        for (x, y, r) in circles:
            
            coord = (x, y)
            
            circle_center_coord.append(coord)
            circle_center_radius.append(r)

        if idx_closest == 0:

            print("Circle marker with radius = {} was detected!\n".format(circle_center_radius[idx_closest]))
        
            '''
            # draw the circle in the output image, then draw a center
            circle_detection_img = cv2.circle(circle_detection_img, circle_center_coord[idx_closest], circle_center_radius[idx_closest], (0, 255, 0), 4)
            circle_detection_img = cv2.circle(circle_detection_img, circle_center_coord[idx_closest], 5, (0, 128, 255), -1)
            '''
            
            # compute the diameter of coin
            diameter_circle = circle_center_radius[idx_closest]*2
            
        
            # mask the detected circle with black color
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            tmp_mask = np.zeros((gray.shape), np.uint8)
            
            #tmp_mask = np.zeros([img_width, img_height], dtype=np.uint8)

            tmp_mask = cv2.circle(tmp_mask, circle_center_coord[idx_closest], circle_center_radius[idx_closest] + 50, (255, 255, 255), -1)

            tmp_mask_binary = cv2.threshold(tmp_mask, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
            
            tmp_mask_binary = cv2.bitwise_not(tmp_mask_binary) 
      
            masked_tmp = cv2.bitwise_and(image.copy(), image.copy(), mask = tmp_mask_binary)
            
            #####################################################
            # save marker part as detection results
            (startX, startY) = circle_center_coord[idx_closest]
            
            sx = startX -r*1 if startX -r*1 > 0 else 0
            sy = startY -r*1 if startY -r*1 > 0 else 0
            
            endX = startX + int(r*1.2) 
            endY = startY + int(r*1.2) 
            
            circle_detection_img = output[sy:endY, sx:endX]
            
            
            ###################################################
            
            # crop ROI region based on the location of circle marker
            offset = 1250
            
            endX = startX + int(r*1.2) + offset
            endY = startY + int(r*1.2) + offset
            
            sx = startX -r*4 if startX -r*4 > 0 else 0
            sy = startY -r*4 if startY -r*4 > 0 else 0

            ROI_region = masked_tmp[sy:endY, sx:endX]
        
        #sticker_crop_img = output
    
    else:
        
        print("No circle was found!\n")
        
        ROI_region = output
        
        masked_tmp = output
        
        diameter_circle = 0
    
    return diameter_circle, ROI_region, circle_detection_img


'''
# Detect stickers in the image
def sticker_detect(img_ori):
    

   

    # load the image, clone it for output, and then convert it to grayscale
    img_rgb = img_ori.copy()
    
    # Convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # Store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    
    #(minVal, maxVal, minLoc, maxLoc) = cv2.minMaxLoc(res)
    
    
    # Specify a threshold 
    threshold = 0.6
    
    if np.amax(res) > threshold:
        
        flag = True
    else:

        flag = False
    
    print(flag)
    

    # Store the coordinates of matched area in a numpy array 
    loc = np.where( res >= threshold)  
    
    if len(loc):
    
        (y,x) = np.unravel_index(res.argmax(), res.shape)
    
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)
    
        #print(y,x)
        
        #print(min_val, max_val, min_loc, max_loc)
        
        
        (startX, startY) = max_loc
        endX = startX + template.shape[0] + 1050 + 110
        endY = startY + template.shape[1] + 1050 + 110
        
        
        # Draw a rectangle around the matched region. 
        for pt in zip(*loc[::-1]): 
            
            sticker_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,255,255), 2)
        
        
        sticker_crop_img = img_rgb[startY:endY, startX:endX]


    return  sticker_crop_img, sticker_overlay
'''




# compute the size and shape info of the foreground
def comp_external_contour(orig, thresh):
    
    #find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    thresh_merged = thresh
    
    if len(contours) > 3:
        
        #####################################################################################
        print("Merging contours...\n")
        
        # using agglomerative clustering algorithm to group contours belonging to same object
        contours_list = [element for element in contours]

        # merge adjacent contours with distance threshold
        gp_contours = agglomerative_cluster(contours_list, threshold_distance = 50.0)

        #test_mask = np.zeros([height, width], dtype="uint8")

        thresh_merged = cv2.drawContours(thresh_merged, gp_contours, -1,255,-1)

        #define result path for labeled images
        #result_img_path = file_path + 'test_mask.png'
        #cv2.imwrite(result_img_path, test_mask)
        
        ################################################################################
        #find contours and get the external one
        contours, hier = cv2.findContours(thresh_merged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    else:
        
        #####################################################################################
        print("No need to merge contours...\n")
    
    
    
    
    
    # sort the contours based on area from largest to smallest
    contours_sorted = sorted(contours, key = cv2.contourArea, reverse = True)
    
    # initialize parameters
    area_c_cmax = 0
   
    img_height, img_width, img_channels = orig.shape
   
    index = 1
    
    trait_img = orig.copy()
    
    area = 0
    
    solidity = 0
    
    compactness = 0
    
    w=h=0
    
    trait_img_bk = orig
    
    
    for index, c in enumerate(contours_sorted):
        
    #for c in contours:
        if index < 1:
    
            #get the bounding rect
            x, y, w, h = cv2.boundingRect(c)
            
            if w>img_width*0.01 and h>img_height*0.01:
                
                roi = trait_img_bk[y:y+h, x:x+w]
                
                print("ROI {} detected ...\n".format(index))
                
                # draw contour
                trait_img = cv2.drawContours(trait_img_bk, contours, -1, (0, 255, 255), 3)
        
                # draw a green rectangle to visualize the bounding rect
                trait_img = cv2.rectangle(trait_img_bk, (x, y), (x+w, y+h), (255, 255, 0), 3)
                
                index+= 1
                
                # get convex hull
                hull = cv2.convexHull(c)
                # draw it in red color
                trait_img = cv2.drawContours(trait_img_bk, [hull], -1, (0, 255, 0), 3)

                '''
                #get the min area rect
                rect = cv2.minAreaRect(c)
                
                (x, y), (minAreaRect_width, minAreaRect_height), angle = rect
                
                box = cv2.boxPoints(rect)
                # convert all coordinates floating point values to int
                box = np.int0(box)
                #draw a red 'nghien' rectangle
                trait_img = cv2.drawContours(trait_img_bk, [box], 0, (0, 255, 255), 3)
                '''
                
                

                
                '''
                # calculate epsilon base on contour's perimeter
                # contour's perimeter is returned by cv2.arcLength
                epsilon = 0.01 * cv2.arcLength(c, True)
                # get approx polygons
                approx = cv2.approxPolyDP(c, epsilon, True)
                # draw approx polygons
                trait_img = cv2.drawContours(orig, [approx], -1, (0, 255, 0), 1)
             
                # hull is convex shape as a polygon
                hull = cv2.convexHull(c)
                trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255))
                '''
                
                '''
                #get the min enclosing circle
                (x, y), radius = cv2.minEnclosingCircle(c)
                # convert all values to int
                center = (int(x), int(y))
                radius = int(radius)
                # and draw the circle in blue
                trait_img = cv2.circle(orig, center, radius, (255, 0, 0), 2)
                '''
                
                area = cv2.contourArea(c)
                print("Leaf area = {0:.2f}... \n".format(area))
                
                
                hull = cv2.convexHull(c)
                hull_area = cv2.contourArea(hull)
                solidity = float(area)/hull_area
                print("solidity = {0:.2f}... \n".format(solidity))
                
                '''
                extLeft = tuple(c[c[:,:,0].argmin()][0])
                extRight = tuple(c[c[:,:,0].argmax()][0])
                extTop = tuple(c[c[:,:,1].argmin()][0])
                extBot = tuple(c[c[:,:,1].argmax()][0])
                
                trait_img = cv2.circle(trait_img_bk, extLeft, 6, (255, 0, 0), -1)
                trait_img = cv2.circle(trait_img_bk, extRight, 6, (255, 0, 0), -1)
                trait_img = cv2.circle(trait_img_bk, extTop, 6, (255, 0, 0), -1)
                trait_img = cv2.circle(trait_img_bk, extBot, 6, (255, 0, 0), -1)
                
                max_width = dist.euclidean(extLeft, extRight)
                max_height = dist.euclidean(extTop, extBot)
                
                if max_width > max_height:
                    trait_img = cv2.line(trait_img_bk, extLeft, extRight, (0,0,255), 2)
                else:
                    trait_img = cv2.line(trait_img_bk, extTop, extBot, (0,0,255), 2)
                '''
                
                max_width = w
                
                max_height = h
                
                longest_dimension = max(max_width, max_height)

                print("Width and height are {0:.2f},{1:.2f}... \n".format(w, h))
        
                compactness = min(max_width,max_height)/max(max_width,max_height)
            
    return trait_img, area, solidity, max_width, max_height, longest_dimension, compactness
    
    
    
    
'''
# individual leaf object segmentation and traits computation
def leaf_traits_computation(orig, labels, result_path, base_name, file_extension):
    
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    leaf_index_rec = []
    contours_rec = []
    area_rec = []
    curv_rec = []
    solidity_rec = []
    major_axis_rec = []
    minor_axis_rec = []
    
    leaf_color_ratio_rec = []
    leaf_color_value_rec = []
    
    box_coord_rec = []
    
    count = 0
    
    

    # curvature computation
    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start = 1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype = "uint8")
        mask[labels == label] = 255
        
        
        #get the medial axis of the contour
        image_skeleton, skeleton = skeleton_bw(mask)

                
        # apply individual object mask
        masked = cv2.bitwise_and(orig, orig, mask = mask)
        
        #individual leaf segmentation and color analysis
        ################################################################################
        mkpath_leaf = os.path.dirname(result_path) + '/leaf' + str(label)
        mkdir(mkpath_leaf)
        result_path_leaf = mkpath_leaf + '/'
        

        #define result path 
        result_img_path = (result_path_leaf + 'leaf_' + str(label) + file_extension)
        cv2.imwrite(result_img_path, masked)
        
        # save skeleton result
        result_file = (result_path_leaf + 'leaf_skeleton_' + str(label) + file_extension)
        cv2.imwrite(result_file, img_as_ubyte(image_skeleton))
        
        
        # color clustering result
        (rgb_colors, counts, hex_colors) = color_region(masked, mask, result_path_leaf, args_num_clusters)
        
        #print("hex_colors = {} {}\n".format(hex_colors, type(hex_colors)))
        
        list_counts = list(counts.values())
        
        #list_hex_colors = list(hex_colors)
        
        #print(type(list_counts))
        
        color_ratio = []
        
        for value_counts, value_hex in zip(list_counts, hex_colors):
            
            #print(percentage(value, np.sum(list_counts)))
            
            color_ratio.append(percentage(value_counts, np.sum(list_counts)))
            
            #print("value_hex = {0}".format(value_hex))
            
            #value_hex.append(value_hex)
            
            
        leaf_color_ratio_rec.append(color_ratio)
        leaf_color_value_rec.append(hex_colors)
        
        # detect contours in the mask and grab the largest one
        #cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        contours, hierarchy = cv2.findContours(mask.copy(),cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        c = max(contours, key = cv2.contourArea)
        
       
        if len(c) >= 10 :

            contours_rec.append(c)
            area_rec.append(cv2.contourArea(c))

        else:
            # optional to "delete" the small contours
            #label_trait = cv2.drawContours(orig, [c], -1, (0, 0, 255), 2)
            print("lack of enough points to fit ellipse")
    
 
    
    contours_rec_sorted = [x for _, x in sorted(zip(area_rec, contours_rec), key=lambda pair: pair[0])]
    
    #cmap = get_cmap(len(contours_rec_sorted)) 
    
    cmap = get_cmap(len(contours_rec_sorted)+1)
    
    
    tracking_backgd = np.zeros(gray.shape, dtype = "uint8")
    #backgd.fill(128)
    
    label_trait = orig
    track_trait = orig
    #clean area record list
    area_rec = []
    #individual leaf traits sorting based on area order 
    ################################################################################
    for i in range(len(contours_rec_sorted)):
        
        c = contours_rec_sorted[i]
        
        #assign unique color value in opencv format
        color_rgb = tuple(reversed(cmap(i)[:len(cmap(i))-1]))
        
        color_rgb = tuple([255*x for x in color_rgb])
        
        
        # get coordinates of bounding box
        
        #(x,y,w,h) = cv2.boundingRect(c)
        
        rect = cv2.minAreaRect(c)
        box = cv2.boxPoints(rect)
        box = np.array(box, dtype="int")
        box_coord_flat = box.flatten()

        box_coord = []
        for item in box_coord_flat:
            box_coord.append(item)
            
        #box_coord_list = list(map(int,box_coord.split()))
        #print(type(box_coord))
        #print("bbox coordinates :{0}".format(box_coord))
        
        
        
        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        #label_trait = cv2.circle(orig, (int(x), int(y)), 3, (0, 255, 0), 2)
        
        #draw filled contour
        #label_trait = cv2.drawContours(orig, [c], -1, color_rgb, -1)
        
        label_trait = cv2.drawContours(orig, [c], -1, color_rgb, 2)
        
        label_trait = cv2.putText(orig, "#{}".format(i+1), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color_rgb, 1)
        #label_trait = cv2.putText(backgd, "#{}".format(i+1), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color_rgb, 1)
        
        #draw mini bounding box
        #label_trait = cv2.drawContours(orig, [box], -1, (0, 255, 0), 2)
        
        #######################################individual leaf curvature computation
        
        #Get rotated bounding ellipse of contour
        ellipse = cv2.fitEllipse(c)
        
        #get paramters of ellipse
        ((xc,yc), (d1,d2), angle) = ellipse
        
        # draw circle at ellipse center
        #label_trait = cv2.ellipse(orig, ellipse, color_rgb, 2)
        #label_trait = cv2.circle(backgd, (int(xc),int(yc)), 10, color_rgb, -1)
        
        track_trait = cv2.circle(tracking_backgd, (int(xc),int(yc)), 5, (255, 255, 255), -1)
        
        
        #draw major radius
        #compute major radius
        rmajor = max(d1,d2)/2
        rminor = min(d1,d2)/2
        
        if angle > 90:
            angle = angle - 90
        else:
            angle = angle + 90
        
        #print(angle)
        
        xtop = xc + math.cos(math.radians(angle))*rmajor
        ytop = yc + math.sin(math.radians(angle))*rmajor
        xbot = xc + math.cos(math.radians(angle+180))*rmajor
        ybot = yc + math.sin(math.radians(angle+180))*rmajor
        
        label_trait = cv2.line(orig, (int(xtop),int(ytop)), (int(xbot),int(ybot)), color_rgb, 1)
                
        c_np = np.vstack(c).squeeze()
        
        x = c_np[:,0]
        y = c_np[:,1]
        
        comp_curv = ComputeCurvature(x, y)
        
        curvature = comp_curv.fit(x, y)
        
        #compute solidity
        solidity = float(cv2.contourArea(c))/cv2.contourArea(cv2.convexHull(c))
        
        #print("solidity = {0:.2f}... \n".format(solidity))
        
        
        #record all traits 
        leaf_index_rec.append(i)
        area_rec.append(cv2.contourArea(c))
        curv_rec.append(curvature)
        
        solidity_rec.append(solidity)
        major_axis_rec.append(rmajor)
        minor_axis_rec.append(rminor)
        
        box_coord_rec.append(box_coord)
    ################################################################################
    
    
    #print('unique labels={0}, len(contours_rec)={1}, len(leaf_index_rec)={2}'.format(np.unique(labels),len(contours_rec),len(leaf_index_rec)))
        
    n_contours = len(contours_rec_sorted)
    
    if n_contours > 0:
        print('average curvature = {0:.2f}\n'.format(sum(curv_rec)/n_contours))
    else:
        n_contours = 1.0
    
    
    #print(leaf_color_ratio_rec)
    
    return sum(curv_rec)/n_contours, label_trait, track_trait, leaf_index_rec, contours_rec, area_rec, curv_rec, solidity_rec, major_axis_rec, minor_axis_rec, leaf_color_ratio_rec, leaf_color_value_rec, box_coord_rec
    
'''

# rgb to hex conversion
def RGB2HEX(color):
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))
    
# rgb to float conversion
def RGB2FLOAT(color):
    return "{:.2f}{:.2f}{:.2f}".format(int(color[0]/255.0), int(color[1]/255.0), int(color[2]/255.0))


# get color map from index
def get_cmap(n, name = 'hsv'):
    '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    RGB color; the keyword argument name must be a standard mpl colormap name.'''
    return plt.get_cmap(name, n)
    
    #import matplotlib.cm
    #return matplotlib.cm.get_cmap(name, n)
    


# cluster colors in the maksed image
def color_region(image, mask, result_path, num_clusters):
    
    # read the image
     #grab image width and height
    (h, w) = image.shape[:2]

    #apply the mask to get the segmentation of plant
    masked_image_ori = cv2.bitwise_and(image, image, mask = mask)
    
    
    # convert to RGB
    image_RGB = cv2.cvtColor(masked_image_ori, cv2.COLOR_BGR2RGB)

    # reshape the image to a 2D array of pixels and 3 color values (RGB)
    pixel_values = image_RGB.reshape((-1, 3))
    
    # convert to float
    pixel_values = np.float32(pixel_values)

    # define stopping criteria
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # number of clusters (K)
    compactness, labels, (centers) = cv2.kmeans(pixel_values, num_clusters, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # convert back to 8 bit values
    centers = np.uint8(centers)

    # flatten the labels array
    labels_flat = labels.flatten()

    # convert all pixels to the color of the centroids
    segmented_image = centers[labels_flat]

    # reshape back to the original image dimension
    segmented_image = segmented_image.reshape(image_RGB.shape)


    segmented_image_BRG = cv2.cvtColor(segmented_image, cv2.COLOR_RGB2BGR)
    
    '''
    if args["debug"] == 1:

        #define result path for labeled images
        result_img_path = result_path + 'masked.png'
        cv2.imwrite(result_img_path, masked_image_ori)
        
        #define result path for labeled images
        result_img_path = result_path + 'clustered.png'
        cv2.imwrite(result_img_path, segmented_image_BRG)

    #define result path for labeled images
    #result_img_path = result_path + 'clustered.png'
    #cv2.imwrite(result_img_path, segmented_image_BRG)
    '''
    '''
    fig = plt.figure()
    ax = Axes3D(fig)        
    for label, pix in zip(labels, segmented_image):
        ax.scatter(pix[0], pix[1], pix[2], color = (centers))
            
    result_file = (result_path + base_name + 'color_cluster_distributation.png')
    plt.savefig(result_file)
    '''
    

    #Show only one chosen cluster 
    #masked_image = np.copy(image)
    masked_image = np.zeros_like(image_RGB)

    # convert to the shape of a vector of pixel values
    masked_image = masked_image.reshape((-1, 3))
    # color (i.e cluster) to render
    #cluster = 2

    cmap = get_cmap(num_clusters + 1)
    


    ####################################################################
    counts = Counter(labels_flat)

    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))
    
    center_colors = centers

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    #rgb_colors = [RGB2FLOAT(ordered_colors[i]) for i in counts.keys()]

    #rgb_colors = [ordered_colors[i] for i in counts.keys()]
    
    rgb_colors = [np.array(ordered_colors[i]).reshape(1, 3) for i in counts.keys()]

    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
   
    if len(index_bkg) > 0:
        
        #remove background color 
        del hex_colors[index_bkg[0]]
        del rgb_colors[index_bkg[0]]
        
        # Using dictionary comprehension to find list 
        # keys having value . 
        delete = [key for key in counts if key == index_bkg[0]] 
      
        # delete the key 
        for key in delete: del counts[key] 
    
    ########################################################################
    #compute color cluster ratio in percentage
    list_counts = list(counts.values())
    
    print("list_counts = {}\n".format(list_counts))
    
    color_ratio = []

    for value_counts, value_hex in zip(list_counts, hex_colors):

        #print(percentage(value, np.sum(list_counts)))

        color_ratio.append(percentage(value_counts, np.sum(list_counts)))

    
   
    return rgb_colors, counts, hex_colors, color_ratio, masked_image_ori, segmented_image_BRG



# normalzie image
def _normalise_image(image, *, image_cmap=None):
    image = img_as_float(image)
    if image.ndim == 2:
        if image_cmap is None:
            image = gray2rgb(image)
        else:
            image = plt.get_cmap(image_cmap)(image)[..., :3]
    return image



# compute the image brightness
def isbright(image_file):
    
    # Set up threshold value for luminous channel, can be adjusted and generalized 
    thresh = 1.5
    
    # Load image file 
    orig = cv2.imread(image_file)
    
    # Make backup image
    image = orig.copy()
    
    # Get file name
    #abs_path = os.path.abspath(image_file)
    
    #filename, file_extension = os.path.splitext(abs_path)
    #base_name = os.path.splitext(os.path.basename(filename))[0]

    image_file_name = Path(image_file).name
    
    # Convert color space to LAB format and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
    
    # Normalize L channel by dividing all pixel values with maximum pixel value
    L = L/np.max(L)
    
    text_bool = "bright" if np.mean(L) < thresh else "dark"
    
    #return image_file_name, np.mean(L), text_bool
    
    #print("np.mean(L) < thresh = {}".format(np.mean(L)))
    
    #return np.mean(L) < thresh
    
    b_bright = 1.0 < thresh
    
    #print("Image brightness: {}\n".foramt(b_bright))
    
    return  b_bright, np.mean(L)
    

# compute the image brightness
def remove_character_string(str_input):
    
    return str_input.replace('#', '')


# compute the mean value of hex colors
def hex_mean_color(color_list):
    
    average_value = (int(remove_character_string(color1), 16) + int(remove_character_string(color2), 16) + int(remove_character_string(color3), 16) + int(remove_character_string(color4), 16))//4
       
    return hex(average_value)




# convert from RGB to LAB space,
# Convert it to LAB color space to access the luminous channel which is independent of colors.
def RGB2LAB(image, mask):
    
    # Make backup image
    image_rgb = image.copy()
    
    # apply object mask
    masked_rgb = cv2.bitwise_and(image_rgb, image_rgb, mask = mask)
    
    # Convert color space to LAB space and extract L channel
    (L, A, B) = cv2.split(cv2.cvtColor(masked_rgb, cv2.COLOR_BGR2LAB))
    

    return masked_rgb, L, A, B
    


#computation of color_difference index
def color_diff_index(ref_color, rgb_colors):
    
    #print(ref_color)
    
    #lab_colors = cv2.cvtColor(rgb_colors, cv2.COLOR_RGB2LAB)
    
    color_diff = []
    
    for index, value in enumerate(rgb_colors): 
        
        curr_color = rgb2lab(value)
        
        # color value from skimage rgb2lab: ranges of Lab values which are: L (0-100), a (-128-127), b (-128-127). 
        # differnt from OpenCV cv2.COLOR_RGB2LAB, need scale to 0~255
        curr_color_scaled = (curr_color + [155, 128, 128]) 

        #color difference in CIE lab space
        diff = deltaE_cie76(ref_color, curr_color_scaled)
        
        diff_value = float(diff)

        #diff = dist.euclidean(std_color_value[1].flatten(), checker_color_value[13].flatten())
        
        color_diff.append(diff_value)
        
        #print("current color value = {}, cluster index = {}, color difference = {}: \n".format(curr_color_scaled, index, diff)) 
    
    #color_diff_index = sum(color_diff) / len(color_diff)
    
        
    return color_diff


# Max RGB filter 
def max_rgb_filter(image):
    
    # split the image into its BGR components
    (B, G, R) = cv2.split(image)
    
    # find the maximum pixel intensity values for each
    # (x, y)-coordinate,, then set all pixel values less
    # than M to zero
    M = np.maximum(np.maximum(R, G), B)
    R[R < M] = 0
    G[G < M] = 0
    B[B < M] = 0
    
    
    # merge the channels back together and return the image
    return cv2.merge([B, G, R])


    

# get file information from the file path using python3
def get_file_info(file_full_path):
    
    p = pathlib.Path(file_full_path)

    filename = p.name

    basename = p.stem

    file_path = p.parent.absolute()

    file_path = os.path.join(file_path, '')
    
    return file_path, filename, basename



# execute script inside program
def execute_script(cmd_line):
    
    try:
        #print(cmd_line)
        #os.system(cmd_line)

        process = subprocess.getoutput(cmd_line)
        
        print(process)
        
        #process = subprocess.Popen(cmd_line, shell = True, stdout = subprocess.PIPE)
        #process.wait()
        #print(process.communicate())
        
    except OSError:
        
        print("Failed ...!\n")



# sort contours based on user defined method
def sort_contours(cnts, method = "left-to-right"):
    
    """sort contours based on user defined method
    
    Inputs: 
    
        cnts: contours extracted from mask image
        
        method: user defined method, default was "left-to-right"
        

    Returns:
    
        sorted_cnts: list of sorted contours 
        
    """   
    
    
    # initialize the reverse flag and sort index
    reverse = False
    i = 0
    
    # handle if we need to sort in reverse
    if method == "right-to-left" or method == "bottom-to-top":
        reverse = True
        
    # handle if we are sorting against the y-coordinate rather than
    # the x-coordinate of the bounding box
    if method == "top-to-bottom" or method == "bottom-to-top":
        i = 1
        
    # construct the list of bounding boxes and sort them from top to bottom
    boundingBoxes = [cv2.boundingRect(c) for c in cnts]
    
    (sorted_cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes), key=lambda b:b[1][i], reverse=reverse))
    
    # return the list of sorted contours and bounding boxes
    return sorted_cnts






# find contours in binary image mask and sort them in left-to-right order
def get_contours(image_thresh):
    
    """find contours in binary image mask and sort them in left-to-right order
    
    Inputs: 
    
        image_thresh: image mask

    Returns:
    
        cnts_sorted:  sorted contours
        
    """
    # find contours in the thresholded image
    cnts = cv2.findContours(image_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    cnts = imutils.grab_contours(cnts)

    # sort the contour based on area size from largest to smallest, and get the first two max contours
    cnts_sorted = sorted(cnts, key = cv2.contourArea, reverse = True)[0:n_object]

    # sort the contours from left to right
    cnts_sorted = sort_contours(cnts_sorted, method = "left-to-right")

   
    print("Sorting {} objects in left-to-right order\n".format(len(cnts_sorted)))
    
    return cnts_sorted




def region_extracted(orig, x, y, w, h):
    
    """compute rect region based on left top corner coordinates and dimension of the region
    
    Inputs: 
    
        orig: image
        
        x, y: left top corner coordinates 
        
        w, h: dimension of the region

    Returns:
    
        roi: region of interest
        
    """   
    roi = orig[y:y+h, x:x+w]
    
    return roi



'''
# Color checker detection based on checker image using threshold in HSV
def color_checker_detection(roi_image_checker, result_path):
    
    
    orig_hsv = cv2.cvtColor(roi_image_checker.copy(), cv2.COLOR_BGR2HSV)

    gray_hsv = cv2.cvtColor(orig_hsv, cv2.COLOR_BGR2GRAY)

    thresh_checker = cv2.threshold(gray_hsv, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #orig = sticker_crop_img.copy()
    
    if np.count_nonzero(thresh_checker) > 0:
    
        thresh_checker = clear_border(thresh_checker)

    

    #color clustering based plant object segmentation
    #thresh_checker = color_cluster_seg(roi_image_checker.copy(), args_colorspace, args_channels, args_num_clusters)
    
    #thresh_checker = color_cluster_seg(roi_image_checker.copy(), str('lab'), str('0'), args_num_clusters)
    
    ####################################################################################
    checker_width_rec = []
    checker_height_rec = []
    
    
    # detect color checker area and get mask of checker area 
    # apply connected component analysis to the thresholded image
    (numLabels, labels, stats, centroids) = cv2.connectedComponentsWithStats(thresh_checker, 4, cv2.CV_32S)
    
    # initialize an output mask 
    mask_checker = np.zeros(gray_hsv.shape, dtype="uint8")

    # loop over the number of unique connected component labels
    for i in range(1, numLabels):

        
        # extract the connected component statistics and centroid for
        # the current label
        x = stats[i, cv2.CC_STAT_LEFT]
        y = stats[i, cv2.CC_STAT_TOP]
        w = stats[i, cv2.CC_STAT_WIDTH]
        h = stats[i, cv2.CC_STAT_HEIGHT]
        
        #print("idx = {}, w = {}, h = {}\n".format(i, w, h))
        
        area = stats[i, cv2.CC_STAT_AREA]
        
        (cX, cY) = centroids[i]
        
        # ensure the width, height, and area are all neither too small
        # nor too big
        keepWidth = w > 30 and w < 1500
        keepHeight = h > 30 and h < 1500
        keepArea = area > 900 and area < 200000
        # ensure the connected component we are examining passes all three tests
        #if all((keepWidth, keepArea)):
        
        text = "{}".format(i)
        
        ratio_w_h = w/h
        
        if keepArea and (ratio_w_h < 1.8):
            # construct a mask for the current connected component and
            # then take the bitwise OR with the mask
            #print("[INFO] keeping connected component '{}'".format(i))
            componentMask = (labels == i).astype("uint8") * 255
            mask_checker = cv2.bitwise_or(mask_checker, componentMask)
            
            checker_width_rec.append(w)
            checker_height_rec.append(h)
            
            #color_checker_detected = cv2.putText(roi_image_checker, text, (cX, cY+0), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
    
    # apply individual object mask
    color_checker_masked = cv2.bitwise_and(roi_image_checker.copy(), roi_image_checker.copy(), mask = mask_checker)
    
    
    
    ################################################################################################
    # find contours in the masked checker image and obtain the color value and name 
    cnts = cv2.findContours(mask_checker.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    cnts = imutils.grab_contours(cnts)
    
    
    # initialize the shape detector and color labeler
    cl = ColorLabeler()
    
    lab_image_checker = cv2.cvtColor(roi_image_checker.copy(), cv2.COLOR_BGR2LAB)
    
    std_color_value = []
    
    checker_color_value = []
    
    color_name_list = []
    
    green_checker_idx = []
    
    # loop over the contours
    for idx, c in enumerate(cnts):
        
        # compute the center of the contour
        M = cv2.moments(c)
        cX = int((M["m10"] / M["m00"]))
        cY = int((M["m01"] / M["m00"]))
        # detect the shape of the contour and label the color

        (color_name, color_value) = cl.label(lab_image_checker, c)
        
        if idx < 1:
            std_color_value.append(cl.lab)
        
        checker_color_value.append(np.array(color_value).reshape((1, 3)))
        
        color_name_list.append(color_name)
        
        
        # draw the contour (x, y)-coordinates  and the name of the color on the image
        c = c.astype("float")
        c = c.astype("int")
        text = "{}".format(idx)
        text_color = "{} {}".format(idx, color_name)
        
        color_checker_detected = cv2.drawContours(roi_image_checker, [c], -1, (0, 255, 0), 2)
        
        color_checker_detected = cv2.putText(roi_image_checker, text_color, (cX - 20, cY + 0), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 255, 255), 1)
        
        
        
        if color_name == "yellow green":
            
            #color_checker_detected = cv2.putText(roi_image_checker, text, (cX-80, cY), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
            
            green_checker_idx.append(idx)
        
        
        
    
    # Convert list of array into list
    std_color_value = [item for t in std_color_value for item in t]
    #std_color_value = [l.tolist() for l in std_color_value]
    
    print("color_name_list = {}\n".format(color_name_list))
    

    
    ################################################################################################

    

    #compute the diagonal path length of each color checker
    avg_width_checker = np.average(checker_width_rec)
    avg_height_checker = np.average(checker_height_rec)
    
    #avg_diagonal_length = 237.0, max_width_checker = 235, max_height_checker = 239
    
    print("checker_width_rec = {}, checker_height_rec = {}\n".format(avg_width_checker, avg_height_checker))
    
    
    return avg_width_checker, avg_height_checker, mask_checker, color_checker_detected, color_checker_masked
    
'''



# apply perspective transform to input image
def Perspective_Transform(roi_image_checker, roi_mask, masked_roi):
    
    
    # find contours in the edged image, keep only the largest
    # ones, and initialize our screen contour
    cnts = cv2.findContours(roi_mask.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    cnts = imutils.grab_contours(cnts)
    
    cnts = sorted(cnts, key = cv2.contourArea, reverse = True)[:10]
    
    checker_Cnt = None
    
    # loop over our contours
    for c in cnts:
        # approximate the contour
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.015 * peri, True)
        # if our approximated contour has four points, then
        # we can assume that we have found our screen
        if len(approx) == 4:
            checker_Cnt = approx
            break
    
    # Apply PerspectiveTransform to the contour on the detected mask
    pts = checker_Cnt.reshape(4, 2)

    # 4 Point OpenCV getPerspective Transform  
    masked_roi_warped = four_point_transform(masked_roi, pts)
    
    
    return masked_roi_warped
    
    


# apply perspective transform to input image based on corner coordinates of a rect
def four_point_transform(image, pts):
    
    # obtain a consistent order of the points and unpack them
    # individually
    rect = order_points(pts)
    (tl, tr, br, bl) = rect
    
    # compute the width of the new image, which will be the
    # maximum distance between bottom-right and bottom-left
    # x-coordiates or the top-right and top-left x-coordinates
    widthA = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
    widthB = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
    maxWidth = max(int(widthA), int(widthB))
    
    # compute the height of the new image, which will be the
    # maximum distance between the top-right and bottom-right
    # y-coordinates or the top-left and bottom-left y-coordinates
    heightA = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
    heightB = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
    maxHeight = max(int(heightA), int(heightB))
    
    # now that we have the dimensions of the new image, construct
    # the set of destination points to obtain a "birds eye view",
    # (i.e. top-down view) of the image, again specifying points
    # in the top-left, top-right, bottom-right, and bottom-left
    # order
    dst = np.array([
        [0, 0],
        [maxWidth - 1, 0],
        [maxWidth - 1, maxHeight - 1],
        [0, maxHeight - 1]], dtype = "float32")
        
    # compute the perspective transform matrix and then apply it
    M = cv2.getPerspectiveTransform(rect, dst)
    warped = cv2.warpPerspective(image, M, (maxWidth, maxHeight))
    
    # return the warped image
    return warped


# sort corner coordinates of a rect
def order_points(pts):
    
    # initialzie a list of coordinates that will be ordered
    # such that the first entry in the list is the top-left,
    # the second entry is the top-right, the third is the
    # bottom-right, and the fourth is the bottom-left
    rect = np.zeros((4, 2), dtype = "float32")
    
    # the top-left point will have the smallest sum, whereas
    # the bottom-right point will have the largest sum
    s = pts.sum(axis = 1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    
    # now, compute the difference between the points, the
    # top-right point will have the smallest difference,
    # whereas the bottom-left will have the largest difference
    diff = np.diff(pts, axis = 1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    
    # return the ordered coordinates
    return rect


# ofind middle points
def midpoint(ptA, ptB):
    return ((ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5)



# get block color value 
def block_color(masked_block, c_max):
    
    cl = ColorLabeler()

    lab_block = cv2.cvtColor(masked_block.copy(), cv2.COLOR_BGR2LAB)

    (color_name, color_value) = cl.label(lab_block, c_max)

    
    
    
    return color_name, color_value



# divide the image into grids and get properities of each block
def grid_seg(input_img, nRows, mCols):
    
    
    img_height, img_width, img_channels = input_img.shape

    # Dimensions of the image
    sizeX = img_width
    sizeY = img_height


    blocks = []
    
    blocks_overlay = input_img.copy()
    
    blocks_bk = input_img.copy()
    
    
    block_color_value = []
    block_width = []
    block_height = []
    block_area = []
            
    grid_area = []
    
    for i in range(0, nRows):
        
        for j in range(0, mCols):
            
            seg = input_img[int(i*sizeY/nRows):int(i*sizeY/nRows) + int(sizeY/nRows),int(j*sizeX/mCols):int(j*sizeX/mCols) + int(sizeX/mCols)]
            
            #blocks_overlay = cv2.putText(blocks_overlay, str("({0}{1})".format(i,j)), (int(j*sizeX/mCols) + + int(sizeY/nRows*0.5), int(i*sizeY/nRows) + int(sizeX/mCols*0.5)), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)
    
            blocks_overlay = cv2.putText(blocks_bk, str(len(blocks)), (int(j*sizeX/mCols) + + int(sizeY/nRows*0.5), int(i*sizeY/nRows) + int(sizeX/mCols*0.5)), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)



            
            h, w, _channels = seg.shape
            
            seg_hsv = cv2.cvtColor(seg.copy(), cv2.COLOR_BGR2HSV)

            seg_gray = cv2.cvtColor(seg_hsv, cv2.COLOR_BGR2GRAY)

            thresh_block = cv2.threshold(seg_gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
            
            #apply the mask to get the segmentation of block
            masked_block = cv2.bitwise_and(seg.copy(), seg.copy(), mask = thresh_block)
            
            # find contours in thresholded image, then grab the largest one
            cnts = cv2.findContours(thresh_block.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            cnts = imutils.grab_contours(cnts)
            
            c_max = max(cnts, key=cv2.contourArea)
            
            area = cv2.contourArea(c_max)
            
            

            (color_name, color_value) = block_color(masked_block, c_max)

            
            
            #print("ID = ({}{}), color_name = {} color_value = {}\n".format(i, j, color_name, color_value))
            
            #cnt_block = cv2.drawContours(masked_block, [c_max], 0, (0, 0, 255), 2)
            
            # compute the rotated bounding box of the contour
            box = cv2.minAreaRect(c_max)
            
            box = cv2.cv.BoxPoints(box) if imutils.is_cv2() else cv2.boxPoints(box)
            
            box = np.array(box, dtype="int")
            
            # order the points in the contour such that they appear
            # in top-left, top-right, bottom-right, and bottom-left
            # order, then draw the outline of the rotated bounding box
            #box = perspective.order_points(box)
            
            cnt_block = cv2.drawContours(masked_block, [box.astype("int")], 0, (0, 255, 0), 1)
            
            # loop over the original points and draw them
            for (x, y) in box:
                cnt_block = cv2.circle(masked_block, (int(x), int(y)), 3, (0, 0, 255), -1)
            
            # unpack the ordered bounding box, then compute the midpoint
            # between the top-left and top-right coordinates, followed by
            # the midpoint between bottom-left and bottom-right coordinates
            (tl, tr, br, bl) = box
            (tltrX, tltrY) = midpoint(tl, tr)
            (blbrX, blbrY) = midpoint(bl, br)
            
            # compute the midpoint between the top-left and top-right points,
            # followed by the midpoint between the top-right and bottom-right
            (tlblX, tlblY) = midpoint(tl, bl)
            (trbrX, trbrY) = midpoint(tr, br)
            
            # draw the midpoints on the image
            cnt_block = cv2.circle(masked_block, (int(tltrX), int(tltrY)), 2, (255, 0, 0), -1)
            cnt_block = cv2.circle(masked_block, (int(blbrX), int(blbrY)), 2, (255, 0, 0), -1)
            cnt_block = cv2.circle(masked_block, (int(tlblX), int(tlblY)), 2, (255, 0, 0), -1)
            cnt_block = cv2.circle(masked_block, (int(trbrX), int(trbrY)), 2, (255, 0, 0), -1)
            
            # draw lines between the midpoints
            cnt_block = cv2.line(masked_block, (int(tltrX), int(tltrY)), (int(blbrX), int(blbrY)), (255, 0, 255), 1)
            cnt_block = cv2.line(masked_block, (int(tlblX), int(tlblY)), (int(trbrX), int(trbrY)), (255, 0, 255), 1)
            
            # compute the Euclidean distance between the midpoints
            dA = dist.euclidean((tltrX, tltrY), (blbrX, blbrY))
            dB = dist.euclidean((tlblX, tlblY), (trbrX, trbrY))
    
            # draw the object sizes on the image
            #cnt_block = cv2.putText(masked_block, "{}".format(dA), (int(tltrX), int(tltrY)), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 1)
            #cnt_block = cv2.putText(masked_block, "{}".format(dB), (int(trbrX), int(trbrY)), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 1)
            
            #text_color = "{}".format(color_name)

            #cnt_block = cv2.putText(masked_block, text_color, (int(tltrX), int(tltrY)), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 0, 255), 1)
            

            
            #print("dA = {}, dB = {} area = {} \n".format(dA,dB, area))



            blocks.append(cnt_block)

            block_color_value.append(color_value)
            
            block_width.append(dB)
            
            block_height.append(dA)
            
            block_area.append(area)
            
            grid_area.append(w*h)


    return blocks, blocks_overlay, block_color_value, block_width, block_height, block_area, grid_area





# Color checker detection
def color_checker_detection(roi_image_checker, result_path):
    
    
    #####################################################
    roi_image = remove(roi_image_checker).copy()

    # extract alpha channel
    alpha = roi_image[:, :, 3]

    # threshold alpha channel to get mask from alpha channel
    roi_mask_ori = cv2.threshold(alpha, 0, 255, cv2.THRESH_BINARY)[1]
    
    
    #########################################################
    k = np.ones((25, 25), np.uint8)  # Define 5x5 kernel
    
    #inv = cv2.bitwise_not(roi_mask_ori)
    
    roi_mask = cv2.erode(roi_mask_ori, k, 1)

    
    ################################################################################


    #apply the mask to get the segmentation of plant
    masked_roi = cv2.bitwise_and(roi_image_checker.copy(), roi_image_checker.copy(), mask = roi_mask)
    
    

    # projective transforamtion 
    masked_roi_warped = Perspective_Transform(roi_image_checker, roi_mask, masked_roi)

    
    
    ###################################################################
    # crop image to remove bottom extra 
    (r_height, r_width) = masked_roi_warped.shape[:2]
    
    #print(r_height, r_width)

    # crop bottom part
    crop_h_ratio = 0.97
    
    start_y = 0  # Starting row (top)
    end_y = int(r_height*crop_h_ratio)    # Ending row (bottom)
    start_x = 0   # Starting column (left)
    end_x = r_width    # Ending column (right)
    
    # Crop the image
    cropped_image = masked_roi_warped[start_y:end_y, start_x:end_x]
    
    color_checker_detected = cropped_image

    
    ####################################################################
    #number of rows
    nRows = 4
    # Number of columns
    mCols = 6
    

    (blocks, blocks_overlay, block_color_value, block_width, block_height, block_area, grid_area) = grid_seg(cropped_image, nRows, mCols)
    
    
    # sort blocks based on area size
    
    index_keep = []
    
    for (i, block) in enumerate(blocks):

        print("ID = {}, color = {}, width = {}, height = {}, area = {}\n".format(i, block_color_value[i], block_width[i], block_height[i], block_area[i]))
        
        if block_area[i] > grid_area[i] *0.5:
            index_keep.append(i)
    
    
    
    selected_block_width = [block_width[i] for i in index_keep]
    
    selected_block_height = [block_height[i] for i in index_keep]
    
    
    print(selected_block_width)
    print(selected_block_height)
    
    if len(selected_block_width) > 0:
        average_width = np.mean(selected_block_width)
        print(f"The average width is: {average_width}")
    else:
        print("The list is empty, cannot calculate average.")
    
    if len(selected_block_height) > 0:
        average_height = np.mean(selected_block_height)
        print(f"The average height is: {average_height}")
    else:
        print("The list is empty, cannot calculate average.")

    
    
    return  average_width, average_height, roi_mask, masked_roi, masked_roi_warped, color_checker_detected, blocks, blocks_overlay
    






def thresh_adjust(thresh, img):

    thresh_adjust = thresh
    
    # get the dimension of the image
    img_height, img_width, img_channels = img.shape
    
    #(h, w) = img.shape[:2]

    #######################################################################################
    # mutiple objects detection
    
    n_object = 2
    
    # find contours in the thresholded image
    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    cnts = imutils.grab_contours(cnts)
    
    # sort the contour based on area size from largest to smallest, and get the first two max contours
    cnts_sorted = sorted(cnts, key = cv2.contourArea, reverse = True)[0:n_object]

    # sort the contours from left to right
    cnts_sorted = sort_contours(cnts_sorted, method = "left-to-right")
    
    
    # initialize background image to draw the contours
    img_overlay_bk = img
    
 
    
    extBot_rec = []

    thresh_tmp = np.zeros(img.shape, np.uint8)
    
    
    # loop over the selected contours
    for idx, c in enumerate(cnts_sorted):
        
        extLeft = tuple(c[c[:,:,0].argmin()][0])
        extRight = tuple(c[c[:,:,0].argmax()][0])
        extTop = tuple(c[c[:,:,1].argmin()][0])
        extBot = tuple(c[c[:,:,1].argmax()][0])
        
        extBot_rec.append(extBot)
        
        print("extBot = {}\n".format(extBot))
        
        print("extBot[0] = {}\n".format(extBot[0]))
        
        print("extBot[1] = {}\n".format(extBot[1]))
        

        offset = 80
        
        if idx < 1:
            # Rectangle parameters
            start_point = (0, extBot[1] - offset)  # (x, y) coordinates of the top-left corner
            end_point = (int(img_width*0.5), img_height)  # (x, y) coordinates of the bottom-right corner
        else:
            
            start_point = (int(img_width*0.5), extBot[1] - offset)  
            end_point = (img_width, img_height)  
        
        print("start_point = {}, end_point ={}\n".format(start_point, end_point))
        
        color = (0, 0, 0)  # 
        thickness = -1  # Line thickness in pixels, filled
        
        # generate individual mask 
        mask_seg = cv2.drawContours(thresh_tmp, [c], -1, (255,255,255), -1)

        # Draw the rectangle
        mask_seg = cv2.rectangle(thresh_tmp, start_point, end_point, color, thickness)


        
        
    # convert the mask image to gray format
    mask_seg_gray = cv2.cvtColor(mask_seg, cv2.COLOR_BGR2GRAY)
    
    # convert mask to binary
    thresh_adjust = cv2.threshold(mask_seg_gray, 127, 255, cv2.THRESH_BINARY)[1]        
        

    print("extBot_rec = {}\n".format(extBot_rec))
    

    return thresh_adjust





# compute all the traits
def extract_traits(image_file, result_path):

    file_size = int(os.path.getsize(image_file)/MBFACTOR)
    
    (file_path, image_file_name, basename) = get_file_info(image_file)
   
    
    # make the folder to store the results
    #current_path = abs_path + '/'
    
    print("Computing traits for image : {0}\n".format(str(image_file_name)))
    
    '''
    # save folder construction
    if (args['output_path']):
        result_path = args['output_path']
    else:
        mkpath = os.path.dirname(abs_path) + '/' + base_name + '/'
        mkdir(mkpath)
        #result_path = mkpath + '/'
        
        result_path = os.path.join(mkpath, '')
       
    print("results_folder: {0}\n".format(str(result_path)))
    
    '''
    
    #########################################################################################
    # check color brightness
    (b_bright, b_value) = isbright(image_file)
    
    # initilize parameters
    area = solidity = max_width = max_height = avg_curv = n_leaves = diameter_circle = compactness = longest_dimension = 0
        

    ################################################################################
    # load image data
    image = cv2.imread(image_file)
    
    # make backup image
    orig = image.copy()
    
    # get the dimension of the image
    img_height, img_width, img_channels = orig.shape

    #source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)

    #QR_data = 0
    
    ################################################################################
    # output image file info
    if image is not None:
        
        print("Plant object segmentation using automatic color clustering method... \n")
        
        print("Image file size: {} MB, brightness: {:.2f}, dimension: {} X {}, channels : {}\n".format(str(file_size), b_value, img_height, img_width, img_channels))
    
    ####################################################
    #Color checker detection
    
    #define color checker region
    x = int(img_width*0.46)
    y = int(img_height*0.056)
    w = int(img_width*0.20)
    h = int(img_height*0.157)

    roi_image_checker = region_extracted(orig, x, y, w, h)
    
    #(avg_width_checker, avg_height_checker, mask_checker, color_checker_detected, color_checker_masked) = color_checker_detection(roi_image_checker, result_path)
    
    (average_width, average_height, roi_mask, masked_roi, masked_roi_warped, color_checker_detected, blocks, blocks_overlay) = color_checker_detection(roi_image_checker, result_path)
    
    
    '''
    ###################################################################
    x = int(img_width*0.16)
    y = int(img_height*0.0)
    w = int(img_width*0.70)
    h = int(img_height*0.5)
    
    checker_region = region_extracted(orig, x, y, w, h)
    
    colorspace_par = str("HSV")
    
    channel_par = str("2")

    #thresh_roichecker = color_cluster_seg(region_extracted(orig, x, y, w, h), colorspace_par, channel_par, 2)
    
    #masked_checker_region = remove(region_extracted(orig, x, y, w, h))
    
    thresh_roichecker = remove(checker_region, only_mask = True)
    
    #apply the mask to get the segmentation of plant
    masked_checker_region = cv2.bitwise_and(checker_region, checker_region, mask = thresh_roichecker)
    
    (avg_width_checker, avg_height_checker, mask_checker, color_checker_detected, color_checker_masked) = color_checker_detection(masked_checker_region, result_path)
    '''

    
    
    if args["debug"] == 1:

        file_extension = '.png'

        mkpath = os.path.dirname(result_path) +'/' + basename 

        mkdir(mkpath)

        image_save_path = mkpath + '/'

        print("image_save_path: {}\n".format(image_save_path))

        # save segmentation result
        #write_image_output(roi_mask, image_save_path, basename, '_roi_mask', file_extension)

        #write_image_output(masked_roi, image_save_path, basename, '_masked_roi', file_extension)

        #write_image_output(masked_roi_warped, image_save_path, basename, '_masked_roi_warped', file_extension)

        #write_image_output(color_checker_detected, image_save_path, basename, '_color_checker_detected', file_extension)
        
        write_image_output(blocks_overlay, image_save_path, basename, '_color_checker', file_extension)
        
        #for (i, block) in enumerate(blocks):

            #result_file = (image_save_path +  str("{:02d}".format(i)) + '.png')

            #cv2.imwrite(result_file, block)
    

    ##########################################################################
    #Plant region detection (defined as ROI_region)
    
    #roi_image = ROI_region.copy()
    
    ROI_region = orig.copy()
    
    #roi_image = orig.copy()
    

    
    ###################################################################################
    # PhotoRoom Remove Background API
    
    if args_ai_switch == 1:
        
        # AI pre-trained model to segment plant object, test function
        roi_image = remove(ROI_region).copy()
        
        #orig = roi_image.copy()

        
        # extract alpha channel
        alpha = roi_image[:, :, 3]

        # threshold alpha channel to get mask from alpha channel
        roi_mask = cv2.threshold(alpha, 0, 255, cv2.THRESH_BINARY)[1]
        
       
        #apply the mask to get the segmentation of plant
        #masked_orig = cv2.bitwise_and(image.copy(), image.copy(), mask = roi_mask)
        
        
        #define result path for labeled images
        result_img_path = result_path + 'roi_masked.png'
        cv2.imwrite(result_img_path, roi_mask)
    
    else:
        
        roi_image = orig.copy()
    
    ######################################################################################
    #orig = roi_image.copy()
    

    #color clustering based plant object segmentation, return plant object mask
    thresh = color_cluster_seg(roi_image, args_colorspace, args_channels, 2)
    
    
    
    thresh = thresh_adjust(thresh, roi_image)
    
    #########################################################################################
    
    if args["debug"] == 1:

        file_extension = '.png'

        mkpath = os.path.dirname(result_path) +'/' + basename 

        mkdir(mkpath)

        image_save_path = mkpath + '/'

        print("image_save_path: {}\n".format(image_save_path))

        # save segmentation result
        write_image_output(thresh, image_save_path, basename, '_thresh', file_extension)
        
        # save segmentation result
        #write_image_output(thresh_test, image_save_path, basename, '_thresh_test', file_extension)
    
    
    
    #######################################################################################
    # mutiple objects detection
    
    n_object = 2
    
    # find contours in the thresholded image
    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    cnts = imutils.grab_contours(cnts)
    
    # sort the contour based on area size from largest to smallest, and get the first two max contours
    cnts_sorted = sorted(cnts, key = cv2.contourArea, reverse = True)[0:n_object]

    # sort the contours from left to right
    cnts_sorted = sort_contours(cnts_sorted, method = "left-to-right")
    
    #print("cv2.contourArea(cnts_sorted[0]), cv2.contourArea(cnts_sorted[1])")
    #print(cv2.contourArea(cnts_sorted[0]), cv2.contourArea(cnts_sorted[1]))
    
    #print("left-to-right")
    #print(len(cnts_sorted))
    

    # initialize variables to record the centers, area of contours
    #center_locX = []
    #center_locY = []
    #cnt_area = [0] * n_object
    
    # initialize empty mask image
    #img_thresh = np.zeros(orig.shape, np.uint8)
    
    # initialize background image to draw the contours
    img_overlay_bk = orig
    
    img_thresh_rec = []
    
    
    

    
    # loop over the selected contours
    for idx, c in enumerate(cnts_sorted):
        
        img_thresh = np.zeros(orig.shape, np.uint8)
        
        extLeft = tuple(c[c[:,:,0].argmin()][0])
        extRight = tuple(c[c[:,:,0].argmax()][0])
        extTop = tuple(c[c[:,:,1].argmin()][0])
        extBot = tuple(c[c[:,:,1].argmax()][0])

        img_overlay = cv2.circle(img_overlay_bk, extLeft, 20, (0, 0, 255), -1)
        img_overlay = cv2.circle(img_overlay_bk, extRight, 20, (0, 0, 255), -1)
        img_overlay = cv2.circle(img_overlay_bk, extTop, 20, (0, 0, 255), -1)
        img_overlay = cv2.circle(img_overlay_bk, extBot, 20, (0, 0, 255), -1)
        
        
        # compute the center of the contour
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        
        # draw the contour and center of the shape on the image
        img_overlay = cv2.drawContours(img_overlay_bk, [c], -1, (0, 255, 0), 2)
        img_overlay = cv2.circle(img_overlay_bk, (cX, cY), 14, (0, 0, 255), -1)
        

        # get the boundning box
        x, y, w, h = cv2.boundingRect(c)
        
        # draw a green rectangle to visualize the bounding rect
        img_overlay = cv2.rectangle(img_overlay_bk, (x, y), (x+w, y+h), (255, 255, 0), 3)

        # get convex hull
        hull = cv2.convexHull(c)
        
        # draw it in red color
        img_overlay = cv2.drawContours(img_overlay_bk, [hull], -1, (0, 255, 0), 3)
        
        
        #center_result = cv2.circle(img_thresh, (cX, cY), 14, (0, 0, 255), -1)
        img_overlay = cv2.putText(img_overlay_bk, "{}".format(idx +1), (cX - 20, cY - 20), cv2.FONT_HERSHEY_SIMPLEX, 5.5, (255, 0, 0), 5)
        
        # generate individual mask 
        mask_seg = cv2.drawContours(img_thresh, [c], -1, (255,255,255), -1)
        
        # convert the mask image to gray format
        mask_seg_gray = cv2.cvtColor(mask_seg, cv2.COLOR_BGR2GRAY)
        
        # convert mask to binary
        mask_seg_bw = cv2.threshold(mask_seg_gray, 127, 255, cv2.THRESH_BINARY)[1]
        
        # store all the masks
        img_thresh_rec.append(mask_seg_bw)
        

    

    
    ########################################################################################

    
    file_name_rec = []

    area_rec = []

    solidity_rec = []

    max_width_rec = []

    max_height_rec = []

    compactness_rec = []

    longest_dimension_rec = []

    n_leaves_rec = []

    hex_colors_rec = []

    color_ratio_rec = []

    color_diff_list_rec = []


    print("{} objects are detected\n".format(len(img_thresh_rec)))
    
    
    for idx, current_thresh in enumerate(img_thresh_rec):
        
        ROI_region = image.copy()

        if cv2.countNonZero(current_thresh) == 0:
            
            print("Image is black\n")
            
            area = solidity = max_width = max_height = longest_dimension = n_leaves = 0
            
            color_ratio = hex_colors = color_diff_list = [0,0,0]
            
        else:
            print("Processing object: {}\n".format(idx+1))
            
            ##########################################################################################################
            # color clustering using pre-defined color cluster value by user
            
            print("number of cluster: {}\n".format(args_num_clusters))
            
            #color clustering of masked image
            (rgb_colors, counts, hex_colors, color_ratio, masked_image_ori, segmented_image_BRG) = color_region(ROI_region, current_thresh, result_path, args_num_clusters)
            
            
            
            ###########################################################################################################
             #compute external contour, shape info  
            (trait_img, area, solidity, max_width, max_height, longest_dimension, compactness) = comp_external_contour(ROI_region, current_thresh)

            
            #############################################################################################################
            # color analysis
            #print("hex_colors = {} {}\n".format(hex_colors, type(hex_colors)))
            
            list_counts = list(counts.values())
            
            #list_hex_colors = list(hex_colors)
            
            #print(type(list_counts))
            
            color_ratio = []
            
            for value_counts, value_hex in zip(list_counts, hex_colors):
                
                #print(percentage(value, np.sum(list_counts)))
                
                color_ratio.append(percentage(value_counts, np.sum(list_counts)))
                
                #print("value_hex = {0}".format(value_hex))
                
                #value_hex.append(value_hex)
            
                
            #color_ratio_rec.append(color_ratio)
            #color_value_rec.append(hex_colors)
            
            #print(rgb_colors)
            #print(color_ratio)
            
            sorted_idx_ratio = np.argsort(color_ratio)

            #reverse the order from accending to descending
            sorted_idx_ratio = sorted_idx_ratio[::-1]


            #sort all lists according to sorted_idx_ratio order
            rgb_colors[:] = [rgb_colors[i] for i in sorted_idx_ratio] 
            color_ratio[:] = [color_ratio[i] for i in sorted_idx_ratio]
            hex_colors[:] = [hex_colors[i] for i in sorted_idx_ratio]
            
            color_name_cluster = []
            ratio_color = []

            cl = ColorLabeler()
            
            for index, (ratio_value, color_value, color_hex) in enumerate(zip(color_ratio, rgb_colors, hex_colors)): 
            
                # validation test color value
                #Skimage rgb2lab outputs 0≤L≤100, −127≤a≤127, −127≤b≤127 . The values are then converted to the destination data type:
                # To convert to opencv (0~255): 8-bit images: L←L∗255/100,a←a+128,b←b+128
                
                #curr_color_lab = rgb2lab([157/255.0, 188/255.0, 64/255.0])
                
                # color convertion between opencv lab data range and Skimage rgb2lab data range
                
                #print(type(color_value))
                
                #print((color_value.shape))
                color_value_reshape = color_value.reshape((1,3))
                
                color_value_float = np.asarray([color_value_reshape[:, 0]/255.0, color_value_reshape[:,1]/255.0, color_value_reshape[:,2]/255.0])
                
                # colorspace teransformation from RGB to LAB 
                curr_color_lab = rgb2lab(color_value_float.flatten())
                
                # +128 avoid the negative numbers when convert the image data to opencv format 
                curr_color_lab_scaled = np.asarray([curr_color_lab[0]*255/100.0, curr_color_lab[1] + 128.0, curr_color_lab[2] + 128.0])
                
                
                print('color_value = {0}, curr_color_lab_scaled = {1}\n'.format(color_value, curr_color_lab_scaled))
                
                color_name = cl.label_c(curr_color_lab_scaled.flatten())
                
                print('Percentage = {0}, rgb_color = {1}, lab_color = {2}, color_name = {3}\n'.format(ratio_value, color_value_reshape, curr_color_lab, color_name))
                
                color_name_cluster.append(color_name)
                
                #ratio_color.append(str(color_name) + ":  "  + str("{:.2f}".format(float(ratio_value)*100) + "%"))
                
                ratio_color.append(str("{:.0f}".format(float(ratio_value)*100) + "%") + "; RGB:[" + "{:.2f}".format(float(color_value_reshape[:,0]/255.0)) + "," + "{:.2f}".format(float(color_value_reshape[:,1]/255.0)) + "," + "{:.2f}".format(float(color_value_reshape[:,2]/255.0)) + "]")
                

            print("color_ratio = {}\n".format(color_ratio))

            #######################################################################
            
            # get reference color from selected color checker
            #ref_color = rgb2lab(np.uint8(np.asarray([[rgb_colors_sticker[0]]])))
            
            ####################################################
            # compute the distance between the current L*a*b* color value in color checker and the mean of the plant surface image in CIE lab space
            
            #print("Detected color checker value in lab: skin = {} foliage = {} purple = {}\n".format(checker_color_value[13], checker_color_value[15], checker_color_value[8]))

            
            #if len(green_checker_idx) > 0:
            
                #ref_color_list = checker_color_value[green_checker_idx[0]]
                
            #else:
                #ref_color_list = [(157, 188, 64)]
            
            
            ref_color_list = [(157, 188, 64)]
            
            color_diff_list = []
            
            for ref_color in ref_color_list:
                
                color_diff_index_value = color_diff_index(ref_color, rgb_colors)
            
                print('color_diff_index_value = {0}\n'.format(color_diff_index_value))
            
                color_diff_list.append(color_diff_index_value)
                
            color_diff_list = np.hstack(color_diff_list)
            


            ###############################################
            
            #accquire medial axis of segmentation mask
            #image_skeleton = medial_axis_image(thresh)
            
                
            image_skeleton, skeleton = skeleton_bw(current_thresh)

            ############################################## leaf number computation

            #min_distance_value = 3
                
            print("min_distance_value = {}\n".format(min_distance_value))
            
            #watershed based leaf area segmentaiton 
            labels = watershed_seg(ROI_region, current_thresh, min_distance_value)
            
            #n_leaves = int(len(np.unique(labels)))
            

            
            #labels = watershed_seg_marker(orig, thresh, min_distance_value, img_marker)
            
            #individual_object_seg(orig, labels, result_path, base_name, file_extension)

            #save watershed result label image
            #Map component labels to hue val
            label_hue = np.uint8(128*labels/np.max(labels))
            #label_hue[labels == largest_label] = np.uint8(15)
            blank_ch = 255*np.ones_like(label_hue)
            labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

            # cvt to BGR for display
            labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)
            
            labeled_img[label_hue==0] = 0
            

            #################################################################
            n_leaves = int(len(np.unique(labels)))
            
            #n_leaves = int(len((leaf_index_rec)))
            
            print('number of leaves = {0}\n'.format(n_leaves))
            

            longest_axis = max(max_width, max_height)
            
            
            #cm_pixel_ratio = diagonal_line_length/avg_diagonal_length
            
            cm_pixel_ratio = diameter_circle
            
            ##################################################################
            
            #print(filename, area, solidity, max_width, max_height, compactness, longest_dimension, n_leaves)
            
            #print(hex_colors, color_ratio, color_diff_list)
            
            #object_name = basename + "_{}".format(idx + 1)
            
            file_name_rec.append(basename + "_{}".format(idx + 1))
            
            area_rec.append(area)
            
            solidity_rec.append(solidity)
            
            max_width_rec.append(max_width)
            
            max_height_rec.append(max_height)
            
            compactness_rec.append(compactness)
            
            longest_dimension_rec.append(longest_dimension)
            
            n_leaves_rec.append(n_leaves)
            
            hex_colors_rec.append(hex_colors)
            
            color_ratio_rec.append(color_ratio)
            
            color_diff_list_rec.append(color_diff_list)
            
            
            
            #####################################################################################
            if args["debug"] == 1:

                
                file_extension = '.png'
                
                mkpath = os.path.dirname(result_path) +'/' + basename 

                mkdir(mkpath)

                image_save_path = mkpath + '/'
                
                
                print("image_save_path: {}\n".format(image_save_path))

                object_basename = basename + "_{}".format(idx + 1) 

                # save segmentation result
                write_image_output(current_thresh, image_save_path, object_basename, '_mask', file_extension)
                
                #write_image_output(roi_image, image_save_path, object_basename, '_plant_region', file_extension)
                
                write_image_output(trait_img, image_save_path, object_basename, '_excontour', file_extension)
                
                write_image_output(img_as_ubyte(image_skeleton), image_save_path, object_basename, '_skeleton', file_extension)

                write_image_output(masked_image_ori, image_save_path, object_basename, '_masked_roi', file_extension)
                
                write_image_output(segmented_image_BRG, image_save_path, object_basename, '_clustered', file_extension)

                #write_image_output(labeled_img, image_save_path, object_basename, '_label', file_extension)

                write_image_output(img_overlay, image_save_path, basename, '_img_overlay', file_extension)
                

                fig_pie = plt.figure(figsize = (8, 6))
                fig_pie = plt.pie(counts.values(), labels = color_ratio, colors = hex_colors)
                result_img_path = image_save_path + object_basename + '_pie_color.png'
                plt.savefig(result_img_path)
            
            #############################################################################################
    

    
    return file_name_rec, area_rec, solidity_rec, max_width_rec, max_height_rec, compactness_rec, longest_dimension_rec, n_leaves_rec, hex_colors_rec, color_ratio_rec, color_diff_list_rec
    

    


# save result files
def write_image_output(imagearray, result_path, base_name, addition, ext):

    # save segmentation result
    result_file = (result_path + base_name + addition + ext)
    
    #print(result_file)
    
    cv2.imwrite(result_file, imagearray)
    
    # check saved file
    if os.path.exists(result_file):
        print("Result file was saved at {0}\n".format(result_file))

    else:
        print("Result file writing failed!\n")
    




# save results as excel file
def write_excel_output(trait_file, result_list):
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
        #sheet_leaf = wb.create_sheet()
        
        #sheet_leaf.delete_rows(2, sheet_leaf.max_row+1) # for entire sheet


    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet_leaf = wb.create_sheet()

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'leaf_area'
        sheet.cell(row = 1, column = 3).value = 'solidity'
        sheet.cell(row = 1, column = 4).value = 'max_width'
        sheet.cell(row = 1, column = 5).value = 'max_height'
        sheet.cell(row = 1, column = 6).value = 'compactness'
        sheet.cell(row = 1, column = 7).value = 'longest_dimension'
        sheet.cell(row = 1, column = 8).value = 'color_cluster_1_hex_value'
        sheet.cell(row = 1, column = 9).value = 'color_cluster_1_ratio'
        sheet.cell(row = 1, column = 10).value = 'color_cluster_1_difference'
        sheet.cell(row = 1, column = 11).value = 'color_cluster_2_hex_value'
        sheet.cell(row = 1, column = 12).value = 'color_cluster_2_ratio'
        sheet.cell(row = 1, column = 13).value = 'color_cluster_2_difference'
        sheet.cell(row = 1, column = 14).value = 'color_cluster_3_hex_value'
        sheet.cell(row = 1, column = 15).value = 'color_cluster_3_ratio'
        sheet.cell(row = 1, column = 16).value = 'color_cluster_3_difference'
        
        #return image_file_name, area, solidity, max_width, max_height, n_leaves, hex_colors, color_ratio, color_diff_list
        
    for row in result_list:
        sheet.append(row)
   

    #save the csv file
    wb.save(trait_file)





if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", dest = "path", type = str, required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", dest = "filetype", type = str, required = False, default='jpg,png', help = "Image file type")
    ap.add_argument("-o", "--output_path", dest = "output_path", type = str, required = False,    help = "result path")
    ap.add_argument('-s', '--color_space', dest = "color_space", type = str, required = False, default ='lab', help='Color space to use: BGR, HSV, Lab, YCrCb (YCC)')
    ap.add_argument('-c', '--channels', dest = "channels", type = str, required = False, default='2', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num_clusters', dest = "num_clusters", type = int, required = False, default = 4,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    ap.add_argument('-min', '--min_size', dest = "min_size", type = int, required = False, default = 35000,  help = 'min size of object to be segmented.')
    ap.add_argument('-max', '--max_size', dest = "max_size", type = int, required = False, default = 1000000,  help = 'max size of object to be segmented.')
    ap.add_argument('-md', '--min_dist', dest = "min_dist", type = int, required = False, default = 35,  help = 'distance threshold of watershed segmentation.')
    ap.add_argument("-da", "--diagonal", dest = "diagonal", type = float, required = False,  default = math.sqrt(2), help = "diagonal line length(cm) of indiviudal color checker module")
    ap.add_argument("-d", '--debug', dest = 'debug', type = int, required = False,  default = 1, help = "Whehter save image results or not, 1 = yes, 0 = no")
    ap.add_argument("-ai", '--ai_switch', dest = 'ai_switch', type = int, required = False,  default = 0, help = "Whehter use AI segmentation or not, 1 = yes, 0 = no")
    
    #ap.add_argument("-cc", "--cue_color", dest = "cue_color", type = int, required = False,  default = 0, help="use color cue to detect plant object")
    #ap.add_argument("-cl", "--cue_loc", dest = "cue_loc", type = int, required = False,  default = 0, help="use location cue to detect plant object")
    #ap.add_argument("-ob", "--out_boundary", dest = "out_boundary", type = int, required = False,  default = 0, help="whether the plant object was out of the image boudary or not, 1 yes, 0 no, default 0")
    
    args = vars(ap.parse_args())
    

    # input path
    file_path = args["path"]
    
    ext = args['filetype'].split(',') if 'filetype' in args else []
    
    patterns = [os.path.join(file_path, f"*.{p}") for p in ext]
    
    files = [f for fs in [glob.glob(pattern) for pattern in patterns] for f in fs]
    

    if len(files) > 0:
        
        # check image file format 
        extension_type = check_file_type(file_path, None)
        
        print("Number of input images: {}, Image format: {}\n".format(len(files), extension_type))
    
    else:
        
        print("Input folder was empty...\n")
        
        sys.exit(1)
        

    #print("Input image format: {}\n".format(extension_type))
    
    #print("Input images: {}\n".format(files))
    
    
    # result path
    result_path = args["output_path"] if args["output_path"] is not None else file_path
    
    result_path = os.path.join(result_path, '')

    # printout result path
    print("Output path: {}\n".format(result_path))
    
    
    ########################################################################
    #parameters
    
    min_size = args['min_size']

    min_distance_value = args['min_dist']
    
    diagonal_line_length = args['min_dist']
    
    num_clusters = args['num_clusters'] 
    
    
    args_colorspace = args['color_space']
    args_channels = args['channels']
    args_num_clusters = args['num_clusters']
    
    args_ai_switch = args['ai_switch']
    

    #accquire image file list
    #image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(files)
    
    

    #print((imgList))
    #global result_path
    
    n_images = len(imgList)
    
    result_list = []
    

    
    
    # loop execute
    for image_id, image in enumerate(imgList):
        
        # get file information
        (file_path, filename, basename) = get_file_info(image)
        
        print("Processing image {} ... \n".format(file_path))
        
        # main pipeline
        (filename_rec, area_rec, solidity_rec, max_width_rec, max_height_rec, compactness_rec, longest_dimension_rec, n_leaves_rec, hex_colors_rec, color_ratio_rec, color_diff_list_rec) = extract_traits(image, file_path)
        
        for idx, (filename, area, solidity, max_width, max_height, compactness, longest_dimension, n_leaves, hex_colors, color_ratio, color_diff_list) in enumerate(zip(filename_rec, area_rec, solidity_rec, max_width_rec, max_height_rec, compactness_rec, longest_dimension_rec, n_leaves_rec, hex_colors_rec, color_ratio_rec, color_diff_list_rec)):
            
            result_list.append([filename, area, solidity, max_width, max_height, compactness, longest_dimension, n_leaves, 
                            str(hex_colors[0]), color_ratio[0], color_diff_list[0], 
                            str(hex_colors[1]), color_ratio[1], color_diff_list[1], 
                            str(hex_colors[2]), color_ratio[2], color_diff_list[2]])
            
    #########################################################################
    #trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    #output in command window in a sum table
 
    table = tabulate(result_list, headers = ['filename', 'area', 'solidity', 'max_width', 'max_height' ,'avg_curv', 'n_leaves', 'cluster 1', 'cluster 2', 'cluster 3', 'cluster 4', 'cluster 1 hex value', 'cluster 2 hex value', 'cluster 3 hex value', 'cluster 4 hex value'], tablefmt = 'orgtbl')

    print(table + "\n")
    
    
    ########################################################################################
    #trait_file = (result_path + str(pathlib.PurePath(result_path).name) + '_trait.xlsx')
    trait_file = (result_path + basename + '_trait.xlsx')

    write_excel_output(trait_file, result_list)

    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
    else:
        print("Error in saving Result file\n")
    
    '''
    #####################################################################################
    # grants read and write access to all result folders
    print("Make result files accessible...\n")

    access_grant = "chmod 777 -R " + result_path 
    
    print(access_grant + '\n')
    
    execute_script(access_grant)
    '''
    
    

    
